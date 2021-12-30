# -*- coding: UTF-8 -*-    
# 作者：zhangHJ
# 日期：2021/11/13  21:07
from abaqus import *
from abaqusConstants import *
from caeModules import *
from odbAccess import *
from visualization import *
import os
import openpyxl
import xlwt
import xlrd
import math
import time
import sys
import logging
import numpy as np

Mdb()
session.journalOptions.setValues(replayGeometry=INDEX, recoverGeometry=INDEX)
session.journalOptions.setValues(replayGeometry=COORDINATE, recoverGeometry=COORDINATE)
session.viewports['Viewport: 1'].assemblyDisplay.geometryOptions.setValues(datumAxes=OFF)

def create_aggregate():
    for i in range(na):
            mySketch = myModel.ConstrainedSketch(name='sphereProfile', sheetSize=200)
            mySketch.ArcByCenterEnds(center=(b[i][0], b[i][1]), direction=CLOCKWISE,
                                     point1=(b[i][0], b[i][1] + b[i][3]), point2=(b[i][0], b[i][1] - b[i][3]))
            mySketch.Line(point1=(b[i][0], b[i][1] + b[i][3]), point2=(b[i][0], b[i][1] - b[i][3]))
            myConstructionLine = mySketch.ConstructionLine(point1=(b[i][0], b[i][1] + b[i][3]),
                                                           point2=(b[i][0], b[i][1] - b[i][3]))
            myBall = myModel.Part(name='sphere' + str(i), dimensionality=THREE_D, type=DEFORMABLE_BODY)
            myPart = myBall.BaseSolidRevolve(angle=360.0, flipRevolveDirection=OFF, sketch=mySketch)
            myModel.rootAssembly.DatumCsysByDefault(CARTESIAN)
            myModel.rootAssembly.Instance(dependent=OFF, name='sphere' + str(i), part=myBall)
            myModel.rootAssembly.translate(instanceList=('sphere' + str(i),), vector=(0.0, 0.0, b[i][2]))
            del myBall
            del myConstructionLine
            del myPart
            del mySketch

# 判断骨料是否相交
def interact_judgement(data_aggregate, point1):
    x1 = point1[0]
    y1 = point1[1]
    z1 = point1[2]
    r1 = point1[3]
    sign = True
    for ii in data_aggregate:
        x2 = ii[0]
        y2 = ii[1]
        z2 = ii[2]
        r2 = ii[3]
        if x2 - x1 < r1 + r2 and y2 - y1 < r1 + r2 and z2 - z1 < r1 + r2:
            # distance calculate
            distance = math.sqrt((x1 - x2) ** 2 + (y1 - y2) ** 2 + (z1 - z2) ** 2)
            if distance < (r1 + r2):
                sign = False
                break
    return sign


def tuple_merge():
    tuple = (a.instances['Part-Base-1'],)
    for i in range(na):
        tuple += (a.instances['sphere{}'.format(i)],)
    return tuple


def del_parts():
    del mdb.models['Model-1'].parts['Part-Base']
    for i in range(na):
        del mdb.models['Model-1'].parts['sphere{}'.format(i)]


def get_node_coord(excel_name):
    logging.info("get_node_coord")
    wb = openpyxl.Workbook()
    ws1 = wb.create_sheet("AllNode", 0)
    ws1.cell(1, 1, "Node")
    ws1.cell(1, 2, "M-NodeCoordx")
    ws1.cell(1, 3, "M-NodeCoordy")
    ws1.cell(1, 4, "M-NodeCoordz")
    for i in range(NN):
        nd = mdb.models['Model-1'].parts['Part-1'].nodes[i]
        ws1.cell(i+2,1,nd.label)
        ws1.cell(i+2,2,nd.coordinates[0])
        ws1.cell(i+2,3,nd.coordinates[1])
        ws1.cell(i+2,4,nd.coordinates[2])
    ws2 = wb.create_sheet("MElement", 1)
    ws2.cell(1, 1, "ELabel")
    ws2.cell(1, 2, "ENode1")
    ws2.cell(1, 3, "ENode2")
    ws2.cell(1, 4, "ENode3")
    ws2.cell(1, 5, "ENode4")
    ws2.cell(1, 6, "x")
    ws2.cell(1, 7, "y")
    ws2.cell(1, 8, "z")
    for j in range(NME):
        label = a1.sets['Set-Matrix'].elements[j].label
        connect = a1.sets['Set-Matrix'].elements[j].connectivity
        ws2.cell(j + 2, 1, label)
        ws2.cell(j + 2, 2, connect[0]+1)
        ws2.cell(j + 2, 3, connect[1]+1)
        ws2.cell(j + 2, 4, connect[2]+1)
        ws2.cell(j + 2, 5, connect[3]+1)
        nmd1 = mdb.models['Model-1'].parts['Part-1'].nodes[connect[0]].coordinates
        nmd2 = mdb.models['Model-1'].parts['Part-1'].nodes[connect[1]].coordinates
        nmd3 = mdb.models['Model-1'].parts['Part-1'].nodes[connect[2]].coordinates
        nmd4 = mdb.models['Model-1'].parts['Part-1'].nodes[connect[3]].coordinates
        ws2.cell(j + 2, 6, 0.25 * (nmd1[0] + nmd2[0] + nmd3[0] + nmd4[0]))
        ws2.cell(j + 2, 7, 0.25 * (nmd1[1] + nmd2[1] + nmd3[1] + nmd4[1]))
        ws2.cell(j + 2, 8, 0.25 * (nmd1[2] + nmd2[2] + nmd3[2] + nmd4[2]))
    ws3 = wb.create_sheet("AElement", 2)
    ws3.cell(1, 1, "ELabel")
    ws3.cell(1, 2, "ENode1")
    ws3.cell(1, 3, "ENode2")
    ws3.cell(1, 4, "ENode3")
    ws3.cell(1, 5, "ENode4")
    ws3.cell(1, 6, "x")
    ws3.cell(1, 7, "y")
    ws3.cell(1, 8, "z")
    for k in range(NAE):
        label = a1.sets['Set-Balls'].elements[k].label
        connect = a1.sets['Set-Balls'].elements[k].connectivity
        ws3.cell(k + 2, 1, label)
        ws3.cell(k + 2, 2, connect[0]+1)
        ws3.cell(k + 2, 3, connect[1]+1)
        ws3.cell(k + 2, 4, connect[2]+1)
        ws3.cell(k + 2, 5, connect[3]+1)
        nmd1 = mdb.models['Model-1'].parts['Part-1'].nodes[connect[0]].coordinates
        nmd2 = mdb.models['Model-1'].parts['Part-1'].nodes[connect[1]].coordinates
        nmd3 = mdb.models['Model-1'].parts['Part-1'].nodes[connect[2]].coordinates
        nmd4 = mdb.models['Model-1'].parts['Part-1'].nodes[connect[3]].coordinates
        ws3.cell(k + 2, 6, 0.25 * (nmd1[0] + nmd2[0] + nmd3[0] + nmd4[0]))
        ws3.cell(k + 2, 7, 0.25 * (nmd1[1] + nmd2[1] + nmd3[1] + nmd4[1]))
        ws3.cell(k + 2, 8, 0.25 * (nmd1[2] + nmd2[2] + nmd3[2] + nmd4[2]))
    sh = wb[wb.sheetnames[3]]
    sh.title = "Initial"
    wb.save(excel_name)
    wb.close()


LOG_FORMAT = "%(asctime)s - %(levelname)s - %(message)s"
DATE_FORMAT = "%m/%d/%Y %H:%M:%S %p"
logging.basicConfig(level=logging.INFO, filename='./log.txt', format=LOG_FORMAT, datefmt=DATE_FORMAT)
# 定义立方体试块边长（mm）
side_length = 0.1
IT = 12.5
# modeling
tb1 = ()
myModel = mdb.Model(name='Model-1')
mysketch_1 = myModel.ConstrainedSketch(name='mysketch_1', sheetSize=200.0)
mysketch_1.rectangle(point1=(0.0, 0.0), point2=(side_length, side_length))
myPart = myModel.Part(name='Part-Base', dimensionality=THREE_D, type=DEFORMABLE_BODY)
myPart.BaseSolidExtrude(sketch=mysketch_1, depth=side_length)
mdb.models['Model-1'].rootAssembly.Instance(dependent=ON, name='Part-Base-1',part=mdb.models['Model-1'].parts['Part-Base'])
# 给基质骨料等赋予材料属性
mdb.models['Model-1'].Material(name='Material-Matrix')
mdb.models['Model-1'].materials['Material-Matrix'].Density(table=((2500.0,),))
mdb.models['Model-1'].materials['Material-Matrix'].Elastic(table=((40e9, 0.2),))
for i in range(80):
    tb1 = tb1+((900+80*(10*i+20.0)/120-4*(pow((10*i+20.0)/120.0,2)), 20.0+10*i),)
mdb.models['Model-1'].materials['Material-Matrix'].SpecificHeat(temperatureDependency=ON, table=tb1)
mdb.models['Model-1'].materials['Material-Matrix'].Expansion(table=((9.1e-06,),))
mdb.models['Model-1'].materials['Material-Matrix'].Conductivity(table=((1.25,),))
mdb.models['Model-1'].materials['Material-Matrix'].conductivity.setValues(temperatureDependency=ON,
                                    table=((2.0, 20.0), (1.6, 300.0), (1.5, 400.0), (1.4, 500.0)))
mdb.models['Model-1'].Material(name='Material-Aggregate')
mdb.models['Model-1'].materials['Material-Aggregate'].Density(table=((3000.0,),))
mdb.models['Model-1'].materials['Material-Aggregate'].Elastic(table=((40e9, 0.2),))
mdb.models['Model-1'].materials['Material-Aggregate'].SpecificHeat(table=((900,),))
mdb.models['Model-1'].materials['Material-Aggregate'].Expansion(table=((8.1e-06,),))
mdb.models['Model-1'].materials['Material-Aggregate'].Conductivity(table=((1.55,),))
file_aggregate = "E:\Abaqus\Code\CycleInputFile\data_aggregate.txt"
b = np.loadtxt(file_aggregate, delimiter=',',dtype=np.float32)
na = len(open(file_aggregate, 'rU').readlines())
create_aggregate()
a = mdb.models['Model-1'].rootAssembly
tuple_merge = tuple_merge()
a.InstanceFromBooleanMerge(name='Part-1', instances=(tuple_merge),
                           keepIntersections=ON, originalInstances=DELETE, domain=GEOMETRY)
del_parts()
mdb.models['Model-1'].HomogeneousSolidSection(name='Section-Matrix', material='Material-Matrix', thickness=None)
mdb.models['Model-1'].HomogeneousSolidSection(name='Section-Agg', material='Material-Aggregate', thickness=None)
p = mdb.models['Model-1'].parts['Part-1']
c = p.cells
cells = c[0:na]
region = p.Set(cells=cells, name='Set-Balls')
p.SectionAssignment(region=region, sectionName='Section-Agg', offset=0.0,
    offsetType=MIDDLE_SURFACE, offsetField='', thicknessAssignment=FROM_SECTION)
cells = c[na:na+1]
region = p.Set(cells=cells, name='Set-Matrix')
p.SectionAssignment(region=region, sectionName='Section-Matrix', offset=0.0,
    offsetType=MIDDLE_SURFACE, offsetField='', thicknessAssignment=FROM_SECTION)
# 分析步设置
mdb.models['Model-1'].HeatTransferStep(name='Step-1', previous='Initial', timePeriod=300, maxNumInc=100000,
                                       initialInc=0.01, minInc=1e-6, maxInc=100, deltmx=10, amplitude=STEP)
# 场输出设置
mdb.models['Model-1'].fieldOutputRequests['F-Output-1'].setValues(variables=('NT', 'TEMP', 'HFL'))
regionDef=mdb.models['Model-1'].rootAssembly.allInstances['Part-1-1'].sets['Set-Matrix']
mdb.models['Model-1'].HistoryOutputRequest(name='H-Output-1',
    createStepName='Step-1', variables=('NT', ), frequency=LAST_INCREMENT,
    region=regionDef, sectionPoints=DEFAULT, rebar=EXCLUDE)
# 绝对零度和stefan常数设置
mdb.models['Model-1'].setValues(absoluteZero=-273.15, stefanBoltzmann=5.67e-8)
# 定义升温幅值
mdb.models['Model-1'].TabularAmplitude(name='Amp-1', timeSpan=STEP, smooth=SOLVER_DEFAULT,
                                       data=((0.0, 20.0), (300, 30)))
# 定义表面热交换和辐射
a = mdb.models['Model-1'].rootAssembly
s1 = a.instances['Part-1-1'].faces
side1Faces1 = s1[0:6]
region = a.Surface(side1Faces=side1Faces1, name='Surf-ALL')
mdb.models['Model-1'].FilmCondition(name='Int-surChange',
                                    createStepName='Step-1', surface=region, definition=EMBEDDED_COEFF,
                                    filmCoeff=12, filmCoeffAmplitude='', sinkTemperature=1.0,
                                    sinkAmplitude='Amp-1', sinkDistributionType=UNIFORM, sinkFieldName='')
mdb.models['Model-1'].RadiationToAmbient(name='Int-surRadiation',createStepName='Step-1', surface=region, radiationType=AMBIENT,
                                         distributionType=UNIFORM, field='', emissivity=0.2, ambientTemperature=1.0,
                                         ambientTemperatureAmp='Amp-1')
# 定义预定义场
c = p.cells
cells = c[0:na+1]
p.Set(cells=cells, name='Set-All')
a = mdb.models['Model-1'].rootAssembly
region = a.instances['Part-1-1'].sets['Set-All']
mdb.models['Model-1'].Temperature(name='Predefined Field-1',
    createStepName='Initial', region=region, distributionType=UNIFORM,
    crossSectionDistribution=CONSTANT_THROUGH_THICKNESS, magnitudes=(IT, ))
# 定义网格属性
elemType1 = mesh.ElemType(elemCode=DC3D8, elemLibrary=STANDARD)
elemType2 = mesh.ElemType(elemCode=DC3D6, elemLibrary=STANDARD)
elemType3 = mesh.ElemType(elemCode=DC3D4, elemLibrary=STANDARD)
c = p.cells
cells = c[0:na+1]
pickedRegions =(cells, )
p.setElementType(regions=pickedRegions, elemTypes=(elemType1, elemType2,elemType3))
pickedRegions = c[0:na+1]
p.setMeshControls(regions=pickedRegions, elemShape=TET, technique=FREE)
p.seedPart(size=0.01, deviationFactor=0.1, minSizeFactor=0.1)
p.generateMesh()
a.regenerate()
a1 = mdb.models['Model-1'].rootAssembly.instances['Part-1-1']
NN = len(a1.nodes.getByBoundingBox(-1, -1, 0, 1, 1, 1))
NE = len(a1.elements.getByBoundingBox(-1, -1, 0, 1, 1, 1))
NME = len(a1.sets['Set-Matrix'].elements)
NAE = len(a1.sets['Set-Balls'].elements)
excel_name = "E:/Abaqus/Code/CycleOutputFile/NNE.xlsx"
# get_node_coord(excel_name)
logging.info("nord_coord done")
mdb.saveAs(pathName='E:/Abaqus/Workpace/Model-Temp01.cae')
logging.info("saved!!")
# done