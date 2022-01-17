# -*- coding: UTF-8 -*-    
# 作者：zhangHJ
# 日期：2021/11/19  12:35
from abaqus import *
from abaqusConstants import *
from caeModules import *
from odbAccess import *
from visualization import *
import os
import openpyxl
import math
import time
import sys
import logging
import numpy as np

"""
	double elemp[EN][]={0.0};	//elemp[][0]-elem number;
								//elemp[][1-2]- 
								//elemp[][3]-the mean element temperature=(elemp[i][4]+elemp[][5])/2;
								//elemp[i][4]-last time step element temperature(oC); 
								//elemp[][5]-element temperature at each time step(oC) 
								//elemp[][6]-
								//elemp[i][7]-specific volume of vapor and liquid water mixture(m3/kg)(before mass)
								//elemp[i][8]-element vapor presure(MPa);
								//elemp[][9,10,11]-damage parameters and the damage angle in the local coordinate-system(D1e,D2e,Q)
								//elemp[][12]-the type of vapor,0-water,1-saturated water,2-data gap,3-superheated steam
                                //-elemp[i][13]-material strength (MPa)
                                //-elemp[i][14]-damage parameter of E_modulus by dehydration/decomposition
                                //-elemp[i][15]-this step initial liquid water mass
                                //-elemp[i][16]-this step initial vapor water mass
                                //-elemp[i][17]-this step final liquid water mass
                                //-elemp[i][18]-this step final vapor water mass
                                //-elemp[i][19]-additional specific heat capacity
                                //-elemp[i][20]-original specific heat capacity in specific temperature

	double elemdh[EN][]={0.0};//elemdh[i][0]-the total amount of moisture transported per unit volume of hcp before this time step(Kg/m3)
								//elemdh[][1]-initial volume fraction of normal C-S-H;
								//elemdh[][2]-initial volume fraction of pozzolanic C-S-H;
								//elemdh[][3]-initial volume fraction of CH;
								//elemdh[][4]-initial volume fraction of AFt;		
								//elemdh[][5]-initial volume fraction of unhydrated cement;
								//elemdh[][6]-initial volume fraction of unhydrated Silica fume;
								//elemdh[i][7]-initial volume fraction of capillary pores;
								//elemdh[][8]-conversion degree of AFt;
								//elemdh[][9]-conversion degree of CH;
								//elemdh[][10]-conversion degree of C-S-H;
                                //elemdh[][11]-porosity increased because of dehydration;
								//elemdh[i][12]-water released by dehydration (mole) per unit volume of hcp(mol/cm3);								
								//elemdh[][13]-volume fraction of normal C-S-H after dehydration of this time step (including the gel pore);
								//elemdh[][14]-volume fraction of pozzolanic C-S-H after dehydration of this time step (including the gel pore);
								//elemdh[][15]-volume fraction of CH after dehydration of this time step;
								//elemdh[][16]-volume fraction of AFt after dehydration of this time step(including the dehydration product);		
								//elemdh[i][17]-volume fraction of capillary pores after dehydration of this time step;
								//elemdh[i][18]-Moisture-Mass1 the mass of liquid and vapor mixture per unit volume of hcp(g/cm3); 
								//elemdh[i][19]-porosity increased by full dehydration(the degree of dehydration=1.0 );
								//elemdh[i][20]-degree of dehydration=porosity increased by dehydration/porosity increased by full dehydration
												//the value is between 0.0~1.0
								//elemdh[i][21]-element critical pore diameter (m)
								//elemdh[i][22]-the volume fraction of residual can not hydrated blender after dehydration
								//elemdh[i][23]-In the end of This step -total Dehydration-MoistureMass
								//elemdh[i][24]-vh-volume fraction of whole product

    double elempe[EN][]={0.0};	//elempe[][1]-level 1 permeability;
								//elempe[][2]-level 2 permeability;
								//elempe[][3]-level 3 permeability-the permeability of hardened cement paste; (m2)
								//elempe[][4]-E_modulus of dehydrated hcp and of aggregate; (MPa)
								//elempe[][5]-element gas permeability taken the slippage flow into account;
								//elempe[i][6]-element relative permeability, the value is between 1 and zero.
								//elempe[i][7]-the element effective gas permeability (taken the relative permeability into account)(m2)
								//elempe[][8]-element poisson's ratio

    double elemvp[EN][]={0.0};  //elemvp[i][1]-element vapor (gas phase) specific volume (m3/kg)
								//elemvp[i][2]-element vapor compresibility factor Z (dimensionless)
								//elemvp[i][3]-dynamic viscosity of vapor (N*S/m2)
								//elemvp[i][4]-element equivalent density for mass transport analysis
								//elemvp[i][5]-element equivalent thermal capacity for mass transport analysis
								//elemvp[i][6]-element equivalent thermal conductivity (permeability) for mass transport analysis
								//elemvp[i][7]-element vapor slippage flow factor b (MPa)0
								//elemvp[i][8]-LW-SDegree 12 == liquid water saturation degree for relative permeability calculation
								//elemvp[i][9]-1/ZRT for the mass migration analysis in the mechanical analysis
								//elemvp[i][10] --element vapor pressure drop because of mass transport (MPa)
								//elemvp[i][11] --vapor density change because of mass transport (kg/m3)
								//elemvp[i][12] --specific volume of vapor and liquid water mixture(m3/kg)(after mass)
								//elemvp[i][13] --the amount of mass change per unit volume of hcp of this time step (kg/m3)
								//elemvp[i][14] --the total amount of mass transported per unit volume of hcp after this time step (kg/m3)
								//elemvp[i][15] --Moisture-SDgree1 == moisture  saturation degree,if the moisture is converted to liquid water,the volume fraction of capillary pores filled with water
								//elemvp[i][16] --expand coefficient for mass modify
								//elemvp[i][17] --density of liquid water in specific temperature
								//elemvp[i][18] --Moisture-Mass2   --> second elemdh[i][18]   
								//elemvp[i][19] --Moisture-SDgree2 -----moisture  saturation degree
								
"""


def interp(y1, y2, x1, x2, x):
    return (y2 - y1) / (x2 - x1) * (x - x1) + y1


def get_increment(file):
    myodb = openOdb(path=file, readOnly=False)
    lastIncrement = len(myodb.steps['Step-1'].frames) - 1
    myodb.close()
    return lastIncrement


def cal_initial(EName):
    with open(condition_mass) as f:
        text = f.readlines()
        sf = float(text[0].split()[1])  # 硅灰替代率
        hydc = float(text[1].split()[1])  # 水泥水化度
        hydsf = float(text[2].split()[1])  # 硅灰的水化度
        wc = float(text[3].split()[1])  # 水灰比
        dsf = float(text[4].split()[1])  # 硅灰密度
        Em = float(text[5].split()[1])  # 基质的弹性模量
        Eag = float(text[6].split()[1])  # 骨料的弹性模量
        Sd = float(text[7].split()[1])  # 初始饱和度
        Dc0 = float(text[8].split()[1])  # 最小孔隙直径
        Dc1 = float(text[9].split()[1])  # 最大孔隙直径
        fvAFt = float(text[10].split()[1])  # 钙矾石体积分数
        fvCH = float(text[11].split()[1])  # CH体积分数
        fvCSH = float(text[12].split()[1])  # CSH体积分数
        inte = float(text[13].split()[1])  # 完成度
        ftm = float(text[14].split()[1])  # 基质抗拉强度
        fta = float(text[15].split()[1])  # 骨料抗拉强度
    # volume fractions of nomal C-S-H
    CSH = 0.68 * fvCSH * hydc * (1.0 - sf) / (wc + 0.32 * (1.0 - sf) + sf / dsf)
    # volume fractions of pozzolanic C-S-H
    pCSH = 3.77 * sf * hydsf / dsf / (wc + 0.32 * (1.0 - sf) + sf / dsf)
    # volume fraction of CH
    CH = (0.68 * fvCH * hydc * (1.0 - sf) - 1.35 * sf * hydsf / dsf) / (wc + 0.32 * (1.0 - sf) + sf / dsf)
    # volume fraction of AFt
    AFt = 0.68 * fvAFt * hydc * (1.0 - sf) / (wc + 0.32 * (1.0 - sf) + sf / dsf)
    # volume fraction of unhydrated cement
    CE = (1.0 - hydc) * 0.32 * (1.0 - sf) / (wc + 0.32 * (1.0 - sf) + sf / dsf)
    # volume fraction of unhydrated Silica fume
    SF = sf * (1.0 - hydsf) / dsf / (wc + 0.32 * (1.0 - sf) + sf / dsf)
    # volume fraction of capillary pores
    CP = 1.0 - CSH - pCSH - CH - AFt - CE - SF
    # increased volume fraction of capillary pores by fully decomposition
    CPF = 0.51633 * AFt + 0.5448 * CH + 0.19 / (0.68 * fvCSH) * CSH + 0.19 * pCSH + 0.2589 * (
            (1.0 - 0.19 / (0.68 * fvCSH)) * CSH + 0.81 * pCSH)
    # when the conversion degree of CSH is 30.7%, the increased volume fraction of capillary pores: for critical pore diameter calculation
    CP2 = 0.51633 * AFt + 0.5448 * CH + (0.19 / (0.68 * fvCSH) * CSH + 0.19 * pCSH + 0.2589 * (
            (1.0 - 0.19 / (0.68 * fvCSH)) * CSH + 0.81 * pCSH)) * 0.307
    # decompostion degree of CP2
    CP2D = CP2 / CPF
    # pore volume fraction increased by fully decomposition of AFt and CH
    CPAFt = 0.51633 * AFt
    CPCH = 0.5448 * CH
    wb = openpyxl.load_workbook(EName)
    sheet = wb.get_sheet_by_name('Initial')
    sheet.cell(1, 1, "F(CSH)")
    sheet.cell(1, 2, "elemdh[i][1]")
    sheet.cell(1, 3, CSH)
    sheet.cell(2, 1, "F(nCSH)")
    sheet.cell(2, 2, "elemdh[i][2]")
    sheet.cell(2, 3, pCSH)
    sheet.cell(3, 1, "F(CH)")
    sheet.cell(3, 2, "elemdh[i][3]")
    sheet.cell(3, 3, CH)
    sheet.cell(4, 1, "F(AFt)")
    sheet.cell(4, 2, "elemdh[i][4]")
    sheet.cell(4, 3, AFt)
    sheet.cell(5, 1, "unhydrated CE")
    sheet.cell(5, 2, "elemdh[i][4]")
    sheet.cell(5, 3, CE)
    sheet.cell(6, 1, "SF")
    sheet.cell(6, 2, "elemdh[i][6]")
    sheet.cell(6, 3, SF)
    sheet.cell(7, 1, "CP")
    sheet.cell(7, 2, "elemdh[i][7]")
    sheet.cell(7, 3, CP)
    sheet.cell(8, 1, "CPF")
    sheet.cell(8, 2, CPF)
    sheet.cell(9, 1, "CPAFt")
    sheet.cell(9, 2, CPAFt)
    sheet.cell(10, 1, "CPCH")
    sheet.cell(10, 2, CPCH)
    sheet.cell(11, 1, "CP2")
    sheet.cell(11, 2, " 600oC")
    sheet.cell(11, 3, CP2)
    sheet.cell(12, 1, "CP2D")
    sheet.cell(12, 2, "decompostion degree of CP2")
    sheet.cell(12, 3, CP2D)
    wb.save(EName)
    wb.close()
    return CSH, pCSH, CH, AFt, CE, SF, CP, CPF, CP2, CP2D, CPAFt, CPCH, sf, hydc, hydsf, wc, dsf, \
           Em, Eag, Sd, Dc0, Dc1, fvAFt, fvCH, fvCSH, inte, ftm, fta


def output_temp(excel, odb_name):
    def calc_temp():
        # 计算Element温度
        logging.info("calc_temp")
        wbs = openpyxl.load_workbook(excel)
        sheet1 = wbs.get_sheet_by_name('AllNode')
        sheet2 = wbs.get_sheet_by_name('MElement')
        sheet3 = wbs.get_sheet_by_name('AElement')
        sheet2.cell(1, 9, "TEMP_Tail")
        sheet3.cell(1, 9, "TEMP_Tail")
        for i in range(NME):
            node_temp = []
            for j in range(2, 6):
                node = sheet2.cell(i + 2, j).value
                node_temp.append(sheet1.cell(node + 1, 5).value)
            sheet2.cell(i + 2, 9, 0.25 * (node_temp[0] + node_temp[1] + node_temp[2] + node_temp[3]))
        for i in range(NAE):
            node_temp = []
            for j in range(2, 6):
                node = sheet3.cell(i + 2, j).value
                node_temp.append(sheet1.cell(node + 1, 5).value)
            sheet3.cell(i + 2, 9, 0.25 * (node_temp[0] + node_temp[1] + node_temp[2] + node_temp[3]))
        wbs.save(excel)
        wbs.close()

    # 输出节点温度
    logging.info("output_temp")
    wb = openpyxl.load_workbook(excel)
    sheet = wb.get_sheet_by_name('AllNode')
    odb = openOdb(path=odb_name)
    lastFrame = odb.steps['Step-1'].frames[-1]  # 创建变量lastFrame，得到载荷步Step-1的最后一帧
    displacement = lastFrame.fieldOutputs['NT11']  # 创建变量displacement，得到最后一帧的温度场数据
    center = odb.rootAssembly.instances['PART-1-1'].nodeSets['SET-ALL']
    dataCollection = displacement.getSubset(region=center)
    sheet.cell(1, 5, "Temp")
    for i in range(NN):
        data = dataCollection.values[i].data
        sheet.cell(i + 2, 5, data)
    wb.save(excel)
    wb.close()
    odb.close()
    calc_temp()


def calc_decomposition(excel0, excel):
    logging.info("calc_decomposition")
    wb1 = openpyxl.load_workbook(excel0)
    wb2 = openpyxl.load_workbook(excel)
    sheet1 = wb1["MElement"]
    if step != 0:
        sheet3 = wb1["Element-VP-PE"]
    sheet2 = wb2["MElement"]
    logging.info("clac begin")
    global MT
    for i in range(NME):
        elemp[i][4] = 20 if step == 0 else sheet1.cell(i + 2, 9).value
        # 计算单元平均温度以用于热分解计算
        elemp[i][5] = sheet2.cell(i + 2, 9).value
        elemp[i][3] = 0.5 * (elemp[i][4] + elemp[i][5])
        MT = max(MT, elemp[i][3])
        # AFt decomposition
        if (elemp[i][3] > 70.0):
            if (elemdh[i][8] < 1.0):
                elemdh[i][8] = elemdh[i][8] + 1.66e4 * math.exp(-7.0965e3 / (elemp[i][3] + 273.15)) * math.sqrt(
                    1.0 - elemdh[i][8]) * TG
            if (elemdh[i][8] > 1.0):
                elemdh[i][8] = 1.0
        # CH decomposition
        if (elemp[i][3] > 438):
            if (elemdh[i][9] == 0.0):
                elemdh[i][9] = 0.004
            if (elemdh[i][9] < 0.5):
                elemdh[i][9] = elemdh[i][9] + 1.516e7 * math.exp(-1.527e4 / (elemp[i][3] + 273.15)) * math.sqrt(
                    elemdh[i][9]) * TG
            elif (elemdh[i][9] >= 0.5 and elemdh[i][9] < 1.0):
                elemdh[i][9] = elemdh[i][9] + 1.516e7 * math.exp(-1.527e4 / (elemp[i][3] + 273.15)) * math.sqrt(
                    1.0 - elemdh[i][9]) * TG
            if (elemdh[i][9] > 1.0):
                elemdh[i][9] = 1.0
        # CSH decomposition
        ra = 1.0 - elemdh[i][10]
        if (elemp[i][3] >= 600.0 and elemp[i][3] <= 700.0):
            if (elemdh[i][10] == 0.0):
                elemdh[i][10] = 0.005
                ra = 1.0 - elemdh[i][10]
            if (elemdh[i][10] > 0.307):
                d600 = 0.0
            else:
                t = 0.0
                for nt in range(30):
                    t1 = t - (-1.0e-6 * pow(t, 3) + 0.0013 * pow(t, 2) - 0.36 * t + 97.72 - ra * 100) / (
                            -3e-6 * pow(t, 2) + 0.0026 * t - 0.36)
                    if (abs((t1 - t) / t1) <= ERRO):
                        break
                    else:
                        t = t1
                t600 = t1
                d600 = (3e-6 * pow(t600, 2) - 0.0026 * t600 + 0.36) / 100
            if (elemdh[i][10] > 0.744):
                d700 = 0.0
            else:
                t = 0.0
                for nt in range(30):
                    t1 = t - (-2.0e-6 * pow(t, 3) + 0.002 * pow(t, 2) - 0.67 * t + 100.61 - ra * 100) / (
                            -6e-6 * pow(t, 2) + 0.004 * t - 0.67)
                    if (abs((t1 - t) / t1) <= ERRO):
                        break
                    else:
                        t = t1
                t700 = t1
                d700 = (6e-6 * pow(t700, 2) - 0.004 * t700 + 0.67) / 100
            if (elemdh[i][10] < 1.0):
                elemdh[i][10] = elemdh[i][10] + interp(d600, d700, 600, 700, elemp[i][3]) * TG / 60
            else:
                elemdh[i][10] = 1.0
        elif (elemp[i][3] > 700.0 and elemp[i][3] <= 800.0):
            if (elemdh[i][10] > 0.744):
                d700 = 0.0
            else:
                t = 0.0
                for nt in range(30):
                    t1 = t - (-2.0e-6 * pow(t, 3) + 0.002 * pow(t, 2) - 0.67 * t + 100.61 - ra * 100) / (
                            -6e-6 * pow(t, 2) + 0.004 * t - 0.67)
                    if (abs((t1 - t) / t1) <= ERRO):
                        break;
                    else:
                        t = t1
                t700 = t1
                d700 = (6e-6 * pow(t700, 2) - 0.004 * t700 + 0.67) / 100.0
            if (elemdh[i][10] >= 0.9625):
                d800 = (1.0e-7 * pow(60.0, 4) - 3.2e-5 * pow(60.0, 3) + 0.0045 * pow(60.0,
                                                                                     2) - 0.26 * 60.0 + 5.47) / 100.0
            else:
                t = 0.0
                for nt in range(30):
                    t1 = t - (-2.0e-8 * pow(t, 5) + 8.0e-6 * pow(t, 4) - 0.0015 * pow(t, 3) + 0.138 * pow(t,
                                                                                                          2) - 5.47 * t + 99.82 - ra * 100) / (
                                 -1.0e-7 * pow(t, 4) + 3.2e-5 * pow(t, 3) - 0.0045 * pow(t, 2) + 0.26 * t - 5.47)
                    if (abs((t1 - t) / t1) <= ERRO):
                        break
                    else:
                        t = t1
                t800 = t1
                d800 = (1.0e-7 * pow(t800, 4) - 3.2e-5 * pow(t800, 3) + 0.0045 * pow(t800,
                                                                                     2) - 0.26 * t800 + 5.47) / 100.0
            if (elemdh[i][10] < 1.0):
                elemdh[i][10] = elemdh[i][10] + interp(d700, d800, 700, 800, elemp[i][3]) * TG / 60
            if (elemdh[i][10] > 1.0):
                elemdh[i][10] = 1.0
        elif (elemp[i][3] > 800.0):
            if (elemdh[i][10] >= 0.9625):
                d800 = (1.0e-7 * pow(60.0, 4) - 3.2e-5 * pow(60.0, 3) + 0.0045 * pow(60.0,
                                                                                     2) - 0.26 * 60.0 + 5.47) / 100.0
            else:
                t = 0.0
                for nt in range(30):
                    t1 = t - (-2.0e-8 * pow(t, 5) + 8.0e-6 * pow(t, 4) - 0.0015 * pow(t, 3) + 0.138 * pow(t,
                                                                                                          2) - 5.47 * t + 99.82 - ra * 100.0) / (
                                 -1.0e-7 * pow(t, 4) + 3.2e-5 * pow(t, 3) - 0.0045 * pow(t, 2) + 0.26 * t - 5.47)
                    if (abs((t1 - t) / t1) <= ERRO):
                        break
                    else:
                        t = t1
                t800 = t1
                d800 = (1.0e-7 * pow(t800, 4) - 3.2e-5 * pow(t800, 3) + 0.0045 * pow(t800,
                                                                                     2) - 0.26 * t800 + 5.47) / 100
            if (elemdh[i][10] < 1.0):
                elemdh[i][10] = elemdh[i][10] + d800 * TG / 60
            if (elemdh[i][10] > 1.0):
                elemdh[i][10] = 1.0
        # ======================= pore volume fraction generated by decomposition and the calculation of dehydration degree =======================
        #  volume fraction of gel pores in conventional C-S-H
        gCSH = 0.19 / (0.68 * fvCSH)
        # solid CSH volume fraction
        vs = ((0.68 * fvCSH - 0.19) * hydc * (1.0 - sf) + 3.0537 * sf * hydsf / dsf) / (wc + 0.32 * (1 - sf) + sf / dsf)
        # solid nomal CSH volume percentage
        a0 = (0.68 * fvCSH - 0.19) * hydc * (1.0 - sf) / (
                (0.68 * fvCSH - 0.19) * hydc * (1.0 - sf) + 3.0537 * sf * hydsf / dsf)
        # solid pazzolanic CSH volume percentage
        b0 = 3.0537 * sf * hydsf / dsf / ((0.68 * fvCSH - 0.19) * hydc * (1.0 - sf) + 3.0537 * sf * hydsf / dsf)
        # the volume fraction of the capillary pore increased by dehydration of CSH
        vpCSH = gCSH / (1.0 - gCSH) * vs * a0 * elemdh[i][10] + 0.2345 * vs * b0 * elemdh[i][10] + 0.2589 * elemdh[i][
            10] * vs
        vpCSH0 = gCSH / (1.0 - gCSH) * vs * a0 + 0.2345 * vs * b0 + 0.2589 * vs
        # porosity increased because of dehydration;
        elemdh[i][11] = 0.51633 * elemdh[i][8] * elemdh[i][4] + 0.5448 * elemdh[i][9] * elemdh[i][3] + vpCSH
        # porosity increased by full dehydration;
        elemdh[i][19] = 0.51633 * elemdh[i][4] + 0.5448 * elemdh[i][3] + vpCSH0
        # dehydration degree
        elemdh[i][20] = elemdh[i][11] / elemdh[i][19]
        # =======================  the water released by dehydration (mole) =======================
        # water released by dehydration (mol/cm3)
        elemdh[i][12] = elemdh[i][11] / 18.0
        elemdh[i][13] = vs * (1.0 - elemdh[i][10]) * a0 / (1.0 - gCSH)
        elemdh[i][14] = vs * (1.0 - elemdh[i][10]) * b0 / 0.81
        elemdh[i][15] = elemdh[i][3] * (1.0 - elemdh[i][9])
        elemdh[i][16] = elemdh[i][4] * (1.0 - elemdh[i][8])
        elemdh[i][17] = elemdh[i][7] + elemdh[i][11]
        elemdh[i][0] = 0 if step == 0 else sheet3.cell(i + 2, 15).value  # 等于上一步的 elemvp[i][14]
        elemdh[i][23] = (elemdh[i][7] * Sd + 18 * elemdh[i][12]) * 1000  # (g/cm3 ---> kg/m3)
        elemdh[i][18] = elemdh[i][23] + elemdh[i][0]  # (kg/m3)
        # unhydrated cement 、silica fume and hydrated product of AFt、CH、CSH
        elemdh[i][22] = elemdh[i][5] + elemdh[i][6] + elemdh[i][4] * elemdh[i][8] * 0.48367 + \
                        elemdh[i][3] * elemdh[i][9] * 0.4552 + vs * elemdh[i][10] * 0.7411
        # vh-volume fraction of whole product
        elemdh[i][24] = elemdh[i][22] + elemdh[i][17] + elemdh[i][16] + elemdh[i][15] + elemdh[i][14] + elemdh[i][13]
        # ======================= E-modulus degradation prediction =======================
        Gm = Em / 2.0 / (1.0 + U)
        Km = Em / (3.0 * (1.0 - 2.0 * U))
        # E-modulus of dehydration products
        EC3A = 164.2 * pow((1.0 - 0.51633), 3.36)
        ECaO = 194.54 * pow((1.0 - 0.5448), 4)
        EnC2S = 140.2 * pow((1.0 - (gCSH + 0.2589 * (1.0 - gCSH))), 4.65)
        EpC2S = 140.2 * pow((1.0 - (0.19 + 0.2589 * (1.0 - 0.19))), 4.65)
        #  volume fraction of dehydration products
        fC3A = elemdh[i][4] * elemdh[i][8]
        fCaO = elemdh[i][3] * elemdh[i][9]
        fnC2S = elemdh[i][1] - elemdh[i][13]
        fpC2S = elemdh[i][2] - elemdh[i][14]
        # the volume fraction of dehydrated phase
        fi = fC3A + fCaO + fnC2S + fpC2S
        # ======================= the E-modulus of dehydrated phase =======================
        if (fi != 0.0):
            Ei = (EC3A * fC3A + ECaO * fCaO + EnC2S * fnC2S + EpC2S * fpC2S) / fi
            Ki = Ei / (3.0 * (1.0 - 2.0 * U))
            Ke0 = Km + (Ki - Km) * fi / (1.0 + (1.0 - fi) * (Ki - Km) / (Km + 4.0 * Gm / 3.0))
            Ee0 = Ke0 * 3.0 * (1.0 - 2.0 * U)
            RE = Ee0 / Em
            elempe[i][4] = Ee0
        else:
            Ke0 = Km
            Ee0 = Em
            RE = Ee0 / Em
            elempe[i][4] = Ee0
        # ======================= permeability prediciton =======================
        # critical pore diameter
        elemdh[i][21] = (Dc1 - Dc0) / CP2D * elemdh[i][20] + Dc0
        # level 1 permeability
        V1 = elemdh[i][14] / (elemdh[i][14] + elemdh[i][5] + elemdh[i][6] + elemdh[i][15] + elemdh[i][16])
        elempe[i][1] = 1.016e-24 * pow((V1 - 0.17), 2)
        # level 2 permeability
        V1 = elemdh[i][13] / (
                elemdh[i][13] + elemdh[i][14] + elemdh[i][5] + elemdh[i][6] + elemdh[i][15] + elemdh[i][16])
        A = 4.8823
        B = (A * (1.0 - V1) - V1) * math.sqrt(elempe[i][1]) + (A * V1 - (1.0 - V1)) * math.sqrt(7e-23)
        C = math.sqrt(7e-23) * math.sqrt(elempe[i][1])
        K21 = pow((B + math.sqrt(pow(B, 2) + 4.0 * A * C)) / (2.0 * A), 2)
        K22 = pow((B - math.sqrt(pow(B, 2) + 4.0 * A * C)) / (2.0 * A), 2)
        if (K21 >= elempe[i][1] and K21 <= 7e-23):
            elempe[i][2] = K21
        else:
            elempe[i][2] = K22
        # level 3 permeability
        V1 = elemdh[i][17]
        Dc = elemdh[i][21]
        Kh = 5.355e-3 * pow(Dc, 2)
        A = 4.5555
        B = (A * (1.0 - V1) - V1) * math.sqrt(elempe[i][2]) + (A * V1 - (1.0 - V1)) * math.sqrt(Kh)
        C = math.sqrt(Kh) * math.sqrt(elempe[i][2])
        K21 = pow((B + math.sqrt(pow(B, 2) + 4.0 * A * C)) / (2.0 * A), 2)
        K22 = pow((B - math.sqrt(pow(B, 2) + 4.0 * A * C)) / (2.0 * A), 2)
        if (K21 >= elempe[i][2] and K21 <= Kh):
            elempe[i][3] = K21
        else:
            elempe[i][3] = K22
    # =======================  output the conversion degree of Aft, CH, and CSH  =======================
    # 计算水化和分解后基质的弹模并输出全部详情
    e0 = ftm / Em
    for i in range(NME):
        elemp[i][14] = elempe[i][4] / Em
        D1 = elemp[i][14]
        x = e0
        for j in range(50):
            x1 = x - (ftm * math.exp(-A * (x - e0)) - D1 * Em * x) / (-A * ftm * math.exp(-A * (x - e0)) - D1 * Em)
            if math.fabs((x1 - x) / x1) <= ERRO:
                break
            else:
                x = x1
        elemp[i][13] = D1 * Em * x1
    sheet = wb2["Element-DH"]
    sheet.cell(1, 1, "ELabel")
    sheet.cell(1, 2, "CD-AFt")  # elemdh[i][8]
    sheet.cell(1, 3, "CD-CH")  # elemdh[i][9]
    sheet.cell(1, 4, "CD-CSH")  # elemdh[i][10]
    sheet.cell(1, 5, "PoIncrease")  # elemdh[i][11]
    sheet.cell(1, 6, "WaterRelease")  # elemdh[i][12]
    sheet.cell(1, 7, "VF-nCSH")  # elemdh[i][13]
    sheet.cell(1, 8, "VF-pCSH")  # elemdh[i][14]
    sheet.cell(1, 9, "VF-CH")  # elemdh[i][15]
    sheet.cell(1, 10, "VF-AFt")  # elemdh[i][16]
    sheet.cell(1, 11, "VF-CP")  # elemdh[i][17]
    sheet.cell(1, 12, "MoisMass1")  # elemdh[i][18]
    sheet.cell(1, 13, "Fullncrease-CP")  # elemdh[i][19]
    sheet.cell(1, 14, "DehyDegree")  # elemdh[i][20]
    sheet.cell(1, 15, "CriticlPore-D")  # elemdh[i][21]
    sheet.cell(1, 16, "VF-RUnhydrated")  # elemdh[i][22]
    sheet.cell(1, 17, "TDehy-Mass")  # elemdh[i][23]
    sheet.cell(1, 18, "VolumeAll")  # elemdh[i][24]
    sheet.cell(1, 19, "TEMP-AVRG")  # elemp[i][3]
    for i in range(NME):
        sheet.cell(i + 2, 1, a1.sets['Set-Matrix'].elements[i].label)
        sheet.cell(i + 2, 2, elemdh[i][8])
        sheet.cell(i + 2, 3, elemdh[i][9])
        sheet.cell(i + 2, 4, elemdh[i][10])
        sheet.cell(i + 2, 5, elemdh[i][11])
        sheet.cell(i + 2, 6, elemdh[i][12])
        sheet.cell(i + 2, 7, elemdh[i][13])
        sheet.cell(i + 2, 8, elemdh[i][14])
        sheet.cell(i + 2, 9, elemdh[i][15])
        sheet.cell(i + 2, 10, elemdh[i][16])
        sheet.cell(i + 2, 11, elemdh[i][17])
        sheet.cell(i + 2, 12, elemdh[i][18])
        sheet.cell(i + 2, 13, elemdh[i][19])
        sheet.cell(i + 2, 14, elemdh[i][20])
        sheet.cell(i + 2, 15, elemdh[i][21])
        sheet.cell(i + 2, 16, elemdh[i][22])
        sheet.cell(i + 2, 17, elemdh[i][23])
        sheet.cell(i + 2, 18, elemdh[i][24])
        sheet.cell(i + 2, 19, elemp[i][3])
    wb2.save(excel)
    wb1.close()
    wb2.close()


def calc_vapor(excel, flag):
    def output_vapor(flag):
        # 输出结果
        logging.info("output_vapor: flag" + str(flag))
        wbb = openpyxl.load_workbook(excel)
        sheet = wbb["MElement"]
        if not flag:
            sheet.cell(1, 10, "SpecificV1")
            sheet.cell(1, 11, "DynamicV1")
            sheet.cell(1, 12, "VaporType1")
            sheet.cell(1, 13, "Vapor1")
            sheet.cell(1, 14, "TEMP_AVRG")
            sheet.cell(1, 15, "LMass1")
            sheet.cell(1, 16, "VMass1")
            for i in range(NME):
                sheet.cell(i + 2, 10, elemp[i][7])
                sheet.cell(i + 2, 11, elemvp[i][3])
                sheet.cell(i + 2, 12, elemp[i][12])
                sheet.cell(i + 2, 13, elemp[i][8])
                sheet.cell(i + 2, 14, elemp[i][3])
                elemp[i][15] = elemdh[i][18] * elemvp[i][8]
                elemp[i][16] = elemdh[i][18] - elemp[i][15]
                sheet.cell(i + 2, 15, elemp[i][15])
                sheet.cell(i + 2, 16, elemp[i][16])
        else:
            sheet.cell(1, 19, "SpecificV2")
            sheet.cell(1, 20, "DynamicV2")
            sheet.cell(1, 21, "VaporType2")
            sheet.cell(1, 22, "Vapor2")
            sheet.cell(1, 23, "LW-SDegree2")
            sheet.cell(1, 24, "M-SDegree2")
            sheet.cell(1, 25, "LMass2")
            sheet.cell(1, 26, "VMass2")
            for i in range(NME):
                sheet.cell(i + 2, 19, elemp[i][7])
                sheet.cell(i + 2, 20, elemvp[i][3])
                sheet.cell(i + 2, 21, elemp[i][12])
                sheet.cell(i + 2, 22, elemp[i][8])
                sheet.cell(i + 2, 23, elemvp[i][8])
                elemvp[i][19] = 1.0 / elemp[i][7] / 1e3
                sheet.cell(i + 2, 24, elemvp[i][19])
                elemp[i][17] = elemdh[i][18] * elemvp[i][8]
                elemp[i][18] = elemdh[i][18] - elemp[i][17]
                sheet.cell(i + 2, 25, elemp[i][17])
                sheet.cell(i + 2, 26, elemp[i][18])
        wbb.save(excel)
        wbb.close()

    logging.info("calc_vapor: flag" + str(flag))
    # 正式计算蒸气压
    # 膨胀系数确定
    get_coefficient()
    # 全局温度判断
    if MT <= 100:
        # 温度小于100度时此时比体积为1.6958
        for i in range(NME):
            elemp[i][7] = 1 / Sd * 1e-3
            elemp[i][8] = 0.1
            elemp[i][12] = 0
            elemvp[i][1] = 1.6958  # 赋予气相比体积
            elemvp[i][3] = 12.02e-6  # 赋予动态粘度
            elemvp[i][8] = elemdh[i][18] / elemdh[i][17] / 1000  # 赋予混合水饱和度(g/cm3 ---> kg/m3)
        output_vapor(False)
        output_vapor(True)
        return
    wb = openpyxl.load_workbook(vapor_table, data_only=True)
    sheet1 = wb["saturatedWater"]
    sheet2 = wb["superheatedSteam"]
    sheet3 = wb["dynamicViscosity"]
    # 计算每行的有效值个数
    count = [0]
    for l in range(2, 21):
        for k in range(2, 34):
            if not sheet2.cell(l, k).value:
                count.append(k - 2)
                break
    for i in range(NME):
        elemp[i][7] = elemdh[i][17] / elemdh[i][18]  # (m3 / kg)
        if elemp[i][3] <= 100:
            elemp[i][8] = 0.1
            elemp[i][12] = 0
            elemvp[i][1] = 1.6958  # 赋予气相比体积
            elemvp[i][3] = 12.02e-6  # 赋予动态粘度
            elemvp[i][8] = elemdh[i][18] / elemdh[i][17] / 1000  # 赋予水饱和度(g/cm3 ---> kg/m3)
            if elemvp[i][8] >= 1.0:
                elemvp[i][8] = 0.999
        else:
            elemp[i][8] = -1.0
            # 考虑饱和蒸汽压表内的蒸汽压
            for j in range(1, 34):
                lvg = interp(sheet1.cell(j + 1, 4).value, sheet1.cell(j + 2, 4).value,
                             sheet1.cell(j + 1, 1).value, sheet1.cell(j + 2, 1).value, elemp[i][3])
                if (sheet1.cell(j + 1, 1).value <= elemp[i][3] <= sheet1.cell(j + 2, 1).value and
                        elemp[i][7] <= lvg):
                    elemp[i][8] = interp(sheet1.cell(j + 1, 2).value, sheet1.cell(j + 2, 2).value,
                                         sheet1.cell(j + 1, 1).value, sheet1.cell(j + 2, 1).value, elemp[i][3])
                    elemp[i][12] = 1.0
                    elemvp[i][1] = lvg
                    elemvp[i][3] = interp(sheet1.cell(j + 1, 5).value, sheet1.cell(j + 2, 5).value,
                                          sheet1.cell(j + 1, 1).value, sheet1.cell(j + 2, 1).value, elemp[i][3])
                    elemvp[i][8] = (elemp[i][7] - elemvp[i][1]) / elemp[i][7] / (1 - elemvp[j][17] * elemvp[i][1])
                    if elemvp[i][8] >= 1.0:
                        elemvp[i][8] = 0.999
                    break
            if elemp[i][8] == -1.0:
                # 匹配过热蒸汽表范围内的单元蒸汽压
                for j in range(1, 21):
                    # 锁定温度以确定该温度下的比体积上限值
                    volume_limit = [0]
                    dv_limit = [0]
                    if sheet2.cell(j + 1, 1).value <= elemp[i][3] <= sheet2.cell(j + 2, 1).value:
                        for k in range(1, count[j] + 1):
                            volume_limit.append(interp(sheet2.cell(j + 1, k + 1).value, sheet2.cell(j + 2, k + 1).value,
                                                       sheet2.cell(j + 1, 1).value, sheet2.cell(j + 2, 1).value,
                                                       elemp[i][3]))
                            dv_limit.append(interp(sheet3.cell(j + 1, k + 1).value, sheet3.cell(j + 2, k + 1).value,
                                                   sheet3.cell(j + 1, 1).value, sheet3.cell(j + 2, 1).value,
                                                   elemp[i][3]))
                        # 赋以过热蒸气压
                        # 比体积的超限处理1
                        if elemp[i][7] >= volume_limit[1]:
                            elemp[i][12] = 3.1
                            elemvp[i][8] = 0
                            elemp[i][7] = volume_limit[1]
                            elemp[i][8] = 0.1
                            elemvp[i][3] = interp(sheet3.cell(j + 1, 2).value, sheet3.cell(j + 2, 2).value,
                                                  sheet3.cell(j + 1, 1).value, sheet3.cell(j + 2, 1).value,
                                                  elemp[i][3]) * 1e-6
                            elemvp[i][1] = elemp[i][7]
                        # 比体积的超限处理2
                        elif elemp[i][3] > 375 and elemp[i][7] <= volume_limit[len(volume_limit) - 1]:
                            elemp[i][12] = 3.2
                            elemvp[i][8] = 0
                            elemp[i][7] = sheet2.cell(j + 1, len(volume_limit)).value
                            elemp[i][8] = sheet2.cell(1, len(volume_limit)).value
                            elemvp[i][3] = interp(sheet3.cell(j + 1, len(volume_limit)).value,
                                                  sheet3.cell(j + 2, len(volume_limit)).value,
                                                  sheet3.cell(j + 1, 1).value, sheet3.cell(j + 2, 1).value,
                                                  elemp[i][3]) * 1e-6
                            elemvp[i][1] = elemp[i][7]
                        # 赋以表内过热蒸气压
                        elif volume_limit[len(volume_limit) - 1] < elemp[i][7] < volume_limit[1]:
                            for n in range(2, len(volume_limit)):
                                if elemp[i][7] >= volume_limit[n]:
                                    elemvp[i][8] = 0
                                    elemp[i][12] = 3
                                    elemp[i][8] = interp(sheet2.cell(1, n).value, sheet2.cell(1, n + 1).value,
                                                         volume_limit[n - 1], volume_limit[n], elemp[i][7])
                                    elemvp[i][3] = interp(dv_limit[n], dv_limit[n - 1], sheet3.cell(1, n).value,
                                                          sheet3.cell(1, n + 1).value, elemp[i][8]) * 1e-6
                                    elemvp[i][1] = elemp[i][7]
                                    break
                        if elemp[i][8] > 35:
                            elemp[i][8] = 35
            # 考虑数据缺失状态
            if elemp[i][8] == -1.0:
                logging.info("match on data gap")
                for k in range(1, 34):
                    if sheet1.cell(k + 1, 1).value <= elemp[i][3] <= sheet1.cell(k + 2, 1).value:
                        # 饱和表中气相比体积临界值
                        vg1 = interp(sheet1.cell(k + 1, 4).value, sheet1.cell(k + 2, 4).value,
                                     sheet1.cell(k + 1, 1).value, sheet1.cell(k + 2, 1).value, elemp[i][3])
                        # 饱和表中蒸汽压临界值
                        vp1 = interp(sheet1.cell(k + 1, 2).value, sheet1.cell(k + 2, 2).value,
                                     sheet1.cell(k + 1, 1).value, sheet1.cell(k + 2, 1).value, elemp[i][3])
                        # 饱和表中动态粘度临界值
                        dv1 = interp(sheet1.cell(k + 1, 5).value, sheet1.cell(k + 2, 5).value,
                                     sheet1.cell(k + 1, 1).value, sheet1.cell(k + 2, 1).value, elemp[i][3])
                        for j in range(1, 11):
                            if sheet2.cell(j + 1, 1).value <= elemp[i][3] <= sheet2.cell(j + 2, 1).value:
                                nn = count[j]
                                # 过热表中气相比体积临界值
                                vg2 = interp(sheet2.cell(j + 1, nn + 1).value, sheet2.cell(j + 2, nn + 1).value,
                                             sheet2.cell(j + 1, 1).value, sheet2.cell(j + 2, 1).value, elemp[i][3])
                                # 过热表中蒸汽压临界值
                                vp2 = sheet2.cell(1, nn + 1).value
                                # 过热表中动态粘度临界值
                                dv2 = interp(sheet3.cell(j + 1, nn + 1).value, sheet3.cell(j + 2, nn + 1).value,
                                             sheet3.cell(j + 1, 1).value, sheet3.cell(j + 2, 1).value, elemp[i][3])
                                elemp[i][12] = 2  # 赋予状态中间数据缺失状
                                elemvp[i][1] = elemp[i][7]  # 赋予比体积
                                elemvp[i][8] = 0.0  # 液态水饱和度
                                elemp[i][8] = interp(vp1, vp2, vg1, vg2, elemp[i][7])
                                elemvp[i][3] = interp(dv1, dv2, vp1, vp2, elemp[i][8]) * 1e-6
                                break
                        break
    wb.close()
    output_vapor(flag)


def get_limit(i):
    wb = openpyxl.load_workbook(vapor_table)
    sheet2 = wb["superheatedSteam"]
    for j in range(1, 21):
        # 锁定温度以确定该温度下的比体积上限值
        if sheet2.cell(j + 1, 1).value <= elemp[i][3] <= sheet2.cell(j + 2, 1).value:
            limit_volume = interp(sheet2.cell(j + 1, 2).value, sheet2.cell(j + 2, 2).value,
                                  sheet2.cell(j + 1, 1).value, sheet2.cell(j + 2, 1).value,
                                  elemp[i][3])
            wb.close()
            return limit_volume


def get_coefficient():
    wb = openpyxl.load_workbook(vapor_table)
    sheet1 = wb["saturatedWater"]
    for i in range(NME):
        for j in range(1, 34):
            # 锁定温度以确定该温度下的比体积上限值
            if sheet1.cell(j + 1, 1).value <= elemp[i][3] <= sheet1.cell(j + 2, 1).value:
                elemvp[i][16] = interp(sheet1.cell(j + 1, 3).value, sheet1.cell(j + 2, 3).value,
                                       sheet1.cell(j + 1, 1).value, sheet1.cell(j + 2, 1).value, elemp[i][3]) / 0.001
                break
            if elemp[i][3] >= 375:
                elemvp[i][16] = 0.003155 / 0.001
                break
            if elemp[i][3] <= 100:
                elemvp[i][16] = 1
                break
    for j in range(NME):
        elemvp[j][17] = 1000.0 / elemvp[j][16]  # 1000 represent regular density of water
    wb.close()


def calc_mass(excel0, excel, massCae, MassOdb, flag):
    def calc_volume():
        wbv = openpyxl.load_workbook(excel)
        wb2 = openpyxl.load_workbook(excel0)
        if step != 0:
            st1 = wb2["Element-VP-PE"]
        sheet2 = wbv["MElement"]
        for i in range(NME):
            if step != 0:
                coefficient1 = st1.cell(i + 2, 17).value
            else:
                coefficient1 = 1
            coefficient2 = elemvp[i][16]
            elemvp[i][10] = sheet2.cell(i + 2, 15).value - sheet2.cell(i + 2, 13).value  # 蒸气压变化值
            elemvp[i][11] = elemvp[i][10] * elemvp[i][9] * 1e3  # 蒸汽密度变化值
            # 计算湿传输（假设膨胀水消失）
            elemvp[i][13] = (1.0 - elemvp[i][8]) * elemdh[i][17] * elemvp[i][11] - (
                    coefficient2 - coefficient1) * elemvp[i][8] * elemdh[i][17] * elemvp[i][17]  # 此分析步中的质量传输
            mm = elemdh[i][18] + elemvp[i][13]  # 对此分析步湿传输后moisture mass计算
            if mm <= 0:
                limit_volume = get_limit(i)
                logging.info("mass transfer is exceeded")
                massNow = elemdh[i][17] / limit_volume
                elemvp[i][13] = massNow - elemdh[i][18]
            elemdh[i][18] = elemdh[i][18] + elemvp[i][13]
            elemvp[i][14] = elemdh[i][0] + elemvp[i][13]  # 此分析步末的累计质量传输
            elemvp[i][15] = 1.0 / elemp[i][7] / elemvp[i][17]  # 此分析步前的累计质量传输
            elemvp[i][18] = elemdh[i][18]
        wbv.save(excel)
        wbv.close()
        wb2.close()

    # 定义计算等效参数步骤
    def calc_equals():
        logging.info("calc_equals")
        for i in range(NME):
            # =============================计算基质的质量传输相关参数===================================
            # 计算压缩性因素Z
            elemvp[i][2] = elemp[i][8] * elemvp[i][1] / 0.4614 / (elemp[i][3] + 273.15) * 1e3
            if elemp[i][8] == 0.1:
                elemvp[i][2] = 1.0
            # 确定相对传导系数
            # 滑移常数
            elemvp[i][7] = math.exp(-0.5818 * math.log(elempe[i][3]) - 19.1213)
            # 考虑汽相滑移的渗透性
            elempe[i][5] = elempe[i][3] * (1.0 + elemvp[i][7] / (elemp[i][8] * 10.0))
            # 相对渗透性系数
            rr = 0.05 - 22.5 * elemdh[i][17]
            elempe[i][6] = pow(10, elemvp[i][8] * rr) - pow(10, rr) * elemvp[i][8]
            # 单元有效气体渗透性
            elempe[i][7] = elempe[i][5] * elempe[i][6]
            elemvp[i][4] = 8.314 / 18  # 质量传输分析的元素等效密度 (R=R`/M) R=0.4614kJ/(kg*K)
            # 质量传输分析的元素等效热容
            elemvp[i][5] = 1.0 / elemvp[i][2] / (elemp[i][3] + 273.15) * elemdh[i][17] * (1.0 - elemvp[i][8])
            elemvp[i][9] = 1.0 / (0.4614 * elemvp[i][2] * (elemp[i][3] + 273.15))  # 1/ZRT
            # 等效热传导率计算
            elemvp[i][6] = elempe[i][7] / elemvp[i][3] / elemvp[i][1] * 800

    # 定义输出步骤
    def output_result(odb_name):
        logging.info("output_mass")
        wb = openpyxl.load_workbook(excel)
        sheet1 = wb["AllNode"]
        sheet2 = wb["MElement"]
        sheet1.cell(1, 6, "VPMass")
        sheet2.cell(1, 15, "VPMass")
        odb = openOdb(path=odb_name)
        last_frame = odb.steps['Step-1'].frames[-1]  # 创建变量lastFrame，得到载荷步Step-1的最后一帧
        displacement = last_frame.fieldOutputs['NT11']  # 创建变量displacement，得到最后一帧的温度场数据
        # 其中Part-cut为装配体中的几何体，SET-为创建的集合
        center = odb.rootAssembly.instances['PART-1-1'].nodeSets['SET-ALL']
        # 创建变量centerDisplacement，得到region center的场数据
        data_collection = displacement.getSubset(region=center)
        for i in range(NN):
            if data_collection.values[i].data <= 0.1:
                sheet1.cell(i + 2, 6, 0.1)
            else:
                sheet1.cell(i + 2, 6, data_collection.values[i].data)
        for i in range(NME):
            node_vapor = []
            for j in range(2, 6):
                node = sheet2.cell(i + 2, j).value
                node_vapor.append(sheet1.cell(node + 1, 6).value)
            sheet2.cell(i + 2, 15, 0.25 * (node_vapor[0] + node_vapor[1] + node_vapor[2] + node_vapor[3]))
        odb.close()
        # 计算湿传输后的比体积
        wb.save(excel)
        calc_volume()
        wb.close()
        output_vppe()

    def output_vppe():
        wb = openpyxl.load_workbook(excel)
        # 输出质量传输的相关参数
        ws1 = wb["Element-VP-PE"]
        ws1.cell(1, 1, "ELabel")
        # output-----elemvp
        ws1.cell(1, 2, "SpecificVg")  # elemvp[i][1]
        ws1.cell(1, 3, "Z")  # elemvp[i][2]
        ws1.cell(1, 4, "Dviscosity")  # elemvp[i][3]
        ws1.cell(1, 5, "EDensity")  # elemvp[i][4]
        ws1.cell(1, 6, "ESpeHeat")  # elemvp[i][5]
        ws1.cell(1, 7, "EConduct")  # elemvp[i][6]
        ws1.cell(1, 8, "SlipFoctor")  # elemvp[i][7]
        ws1.cell(1, 9, "LW-SDegree1")  # elemvp[i][8]
        ws1.cell(1, 10, "1/ZRT")  # elemvp[i][9]
        ws1.cell(1, 11, "VaporDrop")  # elemvp[i][10]
        ws1.cell(1, 12, "VDensityC")  # elemvp[i][11]
        # SpeVolume--AfterMassTransfer list later
        ws1.cell(1, 14, "ThisMChange")  # elemvp[i][13]
        ws1.cell(1, 15, "TotalMTrans")  # elemvp[i][14]
        ws1.cell(1, 16, "M-SDegree1)")  # elemvp[i][15]
        ws1.cell(1, 17, "WaterExpand")  # elemvp[i][16]
        ws1.cell(1, 18, "WDensity")  # elemvp[i][17] density of liquid water in specific temperature
        ws1.cell(1, 19, "MoisMass2")  # elemdh[i][18]===mm
        # output-----elempe
        ws1.cell(1, 21, "LV1P")  # elempe[i][1]
        ws1.cell(1, 22, "LV2P")  # elempe[i][2]
        ws1.cell(1, 23, "LV3P")  # elempe[i][3]
        ws1.cell(1, 24, "EModulus")  # elempe[i][4]
        ws1.cell(1, 25, "S-Permeab")  # elempe[i][5]
        ws1.cell(1, 26, "R-Permeab")  # elempe[i][6]
        ws1.cell(1, 27, "EffectiveP")  # elempe[i][7]
        for i in range(NME):
            # 输出相关信息到指定文件
            ws1.cell(i + 2, 1, a1.sets['Set-Matrix'].elements[i].label)
            ws1.cell(i + 2, 2, elemvp[i][1])
            ws1.cell(i + 2, 3, elemvp[i][2])
            ws1.cell(i + 2, 4, elemvp[i][3])
            ws1.cell(i + 2, 5, elemvp[i][4])
            ws1.cell(i + 2, 6, elemvp[i][5])
            ws1.cell(i + 2, 7, elemvp[i][6])
            ws1.cell(i + 2, 8, elemvp[i][7])
            ws1.cell(i + 2, 9, elemvp[i][8])
            ws1.cell(i + 2, 10, elemvp[i][9])
            ws1.cell(i + 2, 11, elemvp[i][10])
            ws1.cell(i + 2, 12, elemvp[i][11])
            # SpeVolume--AfterMassTransfer list later
            ws1.cell(i + 2, 14, elemvp[i][13])
            ws1.cell(i + 2, 15, elemvp[i][14])
            ws1.cell(i + 2, 16, elemvp[i][15])
            ws1.cell(i + 2, 17, elemvp[i][16])
            ws1.cell(i + 2, 18, elemvp[i][17])
            ws1.cell(i + 2, 19, elemvp[i][18])
            if MT < 100:
                ws1.cell(i + 2, 19, elemdh[i][18])
            # output-----elempe
            ws1.cell(i + 2, 21, elempe[i][1])
            ws1.cell(i + 2, 22, elempe[i][2])
            ws1.cell(i + 2, 23, elempe[i][3])
            ws1.cell(i + 2, 24, elempe[i][4])
            ws1.cell(i + 2, 25, elempe[i][5])
            ws1.cell(i + 2, 26, elempe[i][6])
            ws1.cell(i + 2, 27, elempe[i][7])
        wb.save(excel)

    # 正式计算
    logging.info("calc_mass")
    if MT <= 100:
        logging.info("no mass trans: step-" + str(step + 1))
        output_vppe()
        return
    mdb.Model(name='Model-Mass-1', objectToCopy=mdb.models['Model-1'])
    del mdb.models['Model-Mass-1'].parts['Part-1'].sectionAssignments[1]
    calc_equals()
    logging.info("Give the matrix material properties ")
    # 定义基质的等效属性
    p = mdb.models['Model-Mass-1'].parts['Part-1']
    for i in range(NME):
        materialName = "Material-M-{}".format(i + 1 + NAE)
        sectionName = "Section-M-{}".format(i + 1 + NAE)
        modelMaterial = mdb.models['Model-Mass-1'].Material(name=materialName)
        modelMaterial.Density(table=((elemvp[i][4],),))
        modelMaterial.SpecificHeat(table=((elemvp[i][5],),))
        modelMaterial.Conductivity(table=((elemvp[i][6],),))
        mdb.models['Model-Mass-1'].HomogeneousSolidSection(name=sectionName, material=materialName, thickness=None)
        setName = 'SET-{}'.format(i + 1 + NAE)
        elements = p.elements[NAE + i:NAE + i + 1]
        p.Set(elements=elements, name=setName)
        region = p.sets[setName]
        p.SectionAssignment(region=region, sectionName=sectionName,
                            offset=0.0, offsetType=MIDDLE_SURFACE, offsetField='',
                            thicknessAssignment=FROM_SECTION)
    # =============================输入骨料的质量传输相关参数===================================
    logging.info("Give the aggregate material properties ")
    # 等效热导率
    mdb.models['Model-Mass-1'].materials['Material-Aggregate'].conductivity.setValues(table=((1e-19,),))
    # 等效密度
    mdb.models['Model-Mass-1'].materials['Material-Aggregate'].density.setValues(table=((0.4614,),))
    # 等效比容
    mdb.models['Model-Mass-1'].materials['Material-Aggregate'].specificHeat.setValues(table=((1e2 / (MT + 273.15),),))
    # 撤除相互作用
    mdb.models['Model-Mass-1'].interactions['Int-surChange'].suppress()
    mdb.models['Model-Mass-1'].interactions['Int-surRadiation'].suppress()
    # 定义边界条件
    a = mdb.models['Model-Mass-1'].rootAssembly
    f1 = a.instances['Part-1-1'].faces
    faces1 = f1[0:6]
    region = a.Set(faces=faces1, name='Set-All')
    mdb.models['Model-Mass-1'].TemperatureBC(name='BC-1', createStepName='Step-1',
                                             region=region, fixed=OFF, distributionType=UNIFORM, fieldName='',
                                             magnitude=0.1, amplitude=UNSET)
    # 定义预定义场
    del mdb.models['Model-Mass-1'].predefinedFields['Predefined Field-1']
    # 增量步修改
    # mdb.models['Model-Mass-1'].steps['Step-1'].setValues(minInc=1e-08, deltmx=1.0)
    # 定义映射场
    wb = openpyxl.load_workbook(excel)
    sheet = wb["MElement"]
    vapor_point1 = ()
    for i in range(NME):
        vapor_point1 += ((sheet.cell(i + 2, 6).value, sheet.cell(i + 2, 7).value, sheet.cell(i + 2, 8).value,
                          sheet.cell(i + 2, 13).value),)
    mdb.models['Model-Mass-1'].MappedField(name='AnalyticalField-1',
                                           description='', regionType=POINT, partLevelData=False, localCsys=None,
                                           pointDataFormat=XYZ, fieldDataType=SCALAR, xyzPointData=vapor_point1)
    c1 = a.instances['Part-1-1'].cells
    np = len(c1)
    cells1 = c1[np - 1:np]
    # 为基质赋予初始蒸气压
    region1 = regionToolset.Region(cells=cells1)
    mdb.models['Model-Mass-1'].Temperature(name='Initial-vapor-matrix',
                                           createStepName='Initial', region=region1, distributionType=FIELD,
                                           crossSectionDistribution=CONSTANT_THROUGH_THICKNESS,
                                           field='AnalyticalField-1', magnitudes=(1.0,))
    # 为骨料赋予初始蒸气压
    vapor_point2 = ()
    for n in range(NAN):
        vapor_point2 += ((a1.sets['Set-Balls'].nodes[n].coordinates[0], a1.sets['Set-Balls'].nodes[n].coordinates[1],
                          a1.sets['Set-Balls'].nodes[n].coordinates[2], 0.1),)
    mdb.models['Model-Mass-1'].MappedField(name='AnalyticalField-2',
                                           description='', regionType=POINT, partLevelData=False, localCsys=None,
                                           pointDataFormat=XYZ, fieldDataType=SCALAR, xyzPointData=vapor_point2)
    cells2 = c1[0:np - 1]
    region2 = regionToolset.Region(cells=cells2)
    mdb.models['Model-Mass-1'].Temperature(name='Initial-vapor-balls',
                                           createStepName='Initial', region=region2, distributionType=FIELD,
                                           crossSectionDistribution=CONSTANT_THROUGH_THICKNESS,
                                           field='AnalyticalField-2', magnitudes=(1.0,))
    wb.close()
    logging.info("submit mass job")
    if not flag:
        jobName = "Job-Mass-a-{}".format(step + 1)
    else:
        jobName = "Job-Mass-b-{}".format(step + 1)
    myJob = mdb.Job(name=jobName, model='Model-Mass-1', numCpus=8, numDomains=8, numGPUs=0)
    mdb.jobs[jobName].submit()
    myJob.waitForCompletion()
    mdb.saveAs(pathName=massCae)
    logging.info("mass-clac-done")
    output_result(MassOdb)


def submit_job_a(LastTempCae, LastTempOdb, ThisTempCae):
    logging.info("submit_job")
    global mdb
    if step != 0:
        mdb = openMdb(pathName=LastTempCae)
        # 更改温度计算中的预定义场
        mdb.models['Model-1'].predefinedFields['Predefined Field-1'].setValues(
            distributionType=FROM_FILE, fileName=LastTempOdb, beginStep=1,
            beginIncrement=get_increment(LastTempOdb))
    # 更改分析步
    a1 = mdb.models['Model-1'].rootAssembly.instances['Part-1-1']
    global a1
    mdb.models['Model-1'].steps['Step-1'].setValues(timePeriod=TG)
    # 更改温度幅值
    wb = openpyxl.load_workbook(initial_excel)
    wb2 = openpyxl.load_workbook(excel_base)
    sheet1 = wb['Sheet1']
    sheet2 = wb2['Initial']
    tbt = ()
    count = 1
    for t in range(int(TG / 5.0) + 1):
        tbt = tbt + ((t * 5.0, float(sheet1.cell(t + 4 + step * int(TG / 5.0), 2).value)),)
    for data in tbt:
        sheet2.cell(count, 5, data[0])
        sheet2.cell(count, 6, data[1])
        count += 1
    wb2.save(this_excel_a)
    wb.close()
    wb2.close()
    mdb.models['Model-1'].amplitudes['Amp-1'].setValues(timeSpan=STEP, smooth=SOLVER_DEFAULT, data=tbt)
    jobName = "Job-Temp-a-{}".format(step + 1)
    myJob = mdb.Job(name=jobName, model='Model-1', numCpus=8, numDomains=8, numGPUs=0)
    myJob.submit()
    myJob.waitForCompletion()
    mdb.saveAs(pathName=ThisTempCae)


def upgrade_info(flag):
    for i in range(NME):
        # volume fraction of nomal C-S-H
        elemdh[i][1] = CSH
        # volume fraction of pozzoelanic C-S-H
        elemdh[i][2] = pCSH
        # volume fraction of CH
        elemdh[i][3] = CH
        # volume fraction of AFt
        elemdh[i][4] = AFt
        # volume fraction of unhydrated cement
        elemdh[i][5] = CE
        # volume fraction of unhydrated Silica fume
        elemdh[i][6] = SF
        # volume fraction of capillary pores
        elemdh[i][7] = CP
    if step != 0:
        if not flag:
            wb = openpyxl.load_workbook(last_excel_a)
        else:
            wb = openpyxl.load_workbook(last_excel_b)
        sheet1 = wb['Element-DH']
        for i in range(NME):
            # conversion degree of AFt;
            elemdh[i][8] = sheet1.cell(i + 2, 2).value
            # conversion degree of CH;
            elemdh[i][9] = sheet1.cell(i + 2, 3).value
            # conversion degree of C-S-H;
            elemdh[i][10] = sheet1.cell(i + 2, 4).value
        wb.close()


def node_info(excel):
    def info_output(n1, n2, label, sheet):
        """
        This is a groups style docs.
        Parameters:
          param1 - input column
          param2 - output column
          param3 - output target name
          param3 - input sheet
        """
        data = [[] for n in range(NN)]
        count = 2
        sheet2.cell(count - 1, n2, label)
        for i in range(2, 101):
            value = sheet.cell(i, n1).value
            for j in range(2, 6):
                nodeLabel = sheet1.cell(i, j).value
                data[nodeLabel].append(value)
        for item in data:
            if len(item) == 0:
                result = 0
            else:
                result = sum(item) / len(item)
            sheet2.cell(count, n2, result)
            count = count + 1

    wb = openpyxl.load_workbook(excel)
    sheet1 = wb["MElement"]
    sheet2 = wb["AllNode"]
    sheet3 = wb["Element-VP-PE"]
    sheet4 = wb["Element-DH"]
    info_output(13, 8, "vapor1", sheet1)
    info_output(22, 9, "vapor2", sheet1)
    info_output(12, 10, "MMass1", sheet4)
    info_output(19, 11, "MMass2", sheet3)
    info_output(14, 12, "ThisMLoss", sheet3)
    info_output(15, 13, "TotalMLoss", sheet3)
    wb.save(excel)
    wb.close()


def heat_source(excel_a, excel_b, TempCae):
    def create_model(ThisTempCae, LastTempOdb):
        if step != 0:
            mdb = openMdb(pathName=ThisTempCae)
            # 更改温度计算中的预定义场
            mdb.models['Model-1'].predefinedFields['Predefined Field-1'].setValues(
                distributionType=FROM_FILE, fileName=LastTempOdb, beginStep=1,
                beginIncrement=get_increment(LastTempOdb))
        # 建立新模型
        global mdb
        global a1
        mdb.Model(name='Model-2', objectToCopy=mdb.models['Model-1'])
        a1 = mdb.models['Model-2'].rootAssembly.instances['Part-1-1']
        del mdb.models['Model-2'].parts['Part-1'].sectionAssignments[1]
        # 对上述能量进行作差并除以（考虑湿气吸热前单步内的初始温度和末尾温度的差）得出附加比热
        # 将单元附加比热加到原基质比热上建立新的材料属性
        for i in range(NME):
            p = mdb.models['Model-2'].parts['Part-1']
            materialName = "Material-M{}".format(i + 1 + NAE)
            sectionName = "Section-M{}".format(i + 1 + NAE)
            mdb.models['Model-2'].Material(name=materialName)
            mdb.models['Model-2'].materials[materialName].Density(table=((2500.0,),))
            mdb.models['Model-2'].materials[materialName].SpecificHeat(table=((elemp[i][20],),))
            mdb.models['Model-2'].materials[materialName].Conductivity(temperatureDependency=ON,
                                                                       table=((2.0, 20.0), (1.6, 300.0),
                                                                              (1.5, 400.0), (1.4, 500.0)))
            mdb.models['Model-2'].HomogeneousSolidSection(name=sectionName, material=materialName, thickness=None)
            setName = 'SET-{}'.format(i + 1 + NAE)
            elements = p.elements[NAE + i:NAE + i + 1]
            p.Set(elements=elements, name=setName)
            region = p.sets[setName]
            p.SectionAssignment(region=region, sectionName=sectionName,
                                offset=0.0, offsetType=MIDDLE_SURFACE, offsetField='',
                                thicknessAssignment=FROM_SECTION)

    wb = openpyxl.load_workbook(vapor_table)
    wb2 = openpyxl.load_workbook(excel_a)
    sheet1 = wb["SteamEnergy"]
    st1 = wb2.create_sheet("HeatAbsorb", 0)
    # 未考虑湿气吸热前单步内的初始温度elemp[i][4]和末尾温度elemp[i][5]
    # 1.根据excel结果确定单位体积基质液态水和气态水的质量
    # 2.根据饱和蒸汽表中能量确定单步内初始温度和末尾温度的能量变化值
    st1.cell(1, 1, "uf1")
    st1.cell(1, 2, "ug1")
    st1.cell(1, 3, "uf2")
    st1.cell(1, 4, "ug2")
    st1.cell(1, 5, "energy1")
    st1.cell(1, 6, "energy2")
    st1.cell(1, 7, "addSpeHeat")
    st1.cell(1, 8, "SpeHeat")
    for i in range(NME):
        # a.差值确定初始液态水和气态水单位质量的能量
        # 初始化
        uf1, ug1, uf2, ug2, energy1, energy2 = 0, 0, 0, 0, 0, 0
        for j in range(3, 51):
            if sheet1.cell(j, 1).value <= elemp[i][4] <= sheet1.cell(j + 1, 1).value:
                uf1 = interp(sheet1.cell(j, 5).value, sheet1.cell(j + 1, 5).value, sheet1.cell(j, 1).value,
                             sheet1.cell(j + 1, 1).value, elemp[i][4])
                ug1 = interp(sheet1.cell(j, 6).value, sheet1.cell(j + 1, 6).value, sheet1.cell(j, 1).value,
                             sheet1.cell(j + 1, 1).value, elemp[i][4])
                # 确定初始湿气的总能量
                energy1 = elemp[i][15] * uf1 + elemp[i][16] * ug1
                break
        for k in range(3, 51):
            if sheet1.cell(k, 1).value <= elemp[i][5] <= sheet1.cell(k + 1, 1).value:
                uf2 = interp(sheet1.cell(k, 5).value, sheet1.cell(k + 1, 5).value, sheet1.cell(k, 1).value,
                             sheet1.cell(k + 1, 1).value, elemp[i][5])
                ug2 = interp(sheet1.cell(k, 6).value, sheet1.cell(k + 1, 6).value, sheet1.cell(k, 1).value,
                             sheet1.cell(k + 1, 1).value, elemp[i][5])
                # 确定末尾湿气的总能量
                energy2 = elemp[i][17] * uf2 + elemp[i][18] * ug2 + abs(elemvp[i][13]) * 0.5 * (ug1 + ug2)
                break
        if energy2 - energy1 > 0:
            elemp[i][19] = (energy2 - energy1) / (elemp[i][5] - elemp[i][4])
        elemp[i][20] = 900 + 80 * elemp[i][3] / 120 - 4 * pow((elemp[i][3] / 120), 2) + elemp[i][19]
        st1.cell(i + 2, 1, uf1)
        st1.cell(i + 2, 2, ug1)
        st1.cell(i + 2, 3, uf2)
        st1.cell(i + 2, 4, ug2)
        st1.cell(i + 2, 5, energy1)
        st1.cell(i + 2, 6, energy2)
        st1.cell(i + 2, 6, energy2)
        st1.cell(i + 2, 6, energy2)
        st1.cell(i + 2, 7, elemp[i][19])
        st1.cell(i + 2, 8, elemp[i][20])
    wb2.save(excel_b)
    wb2.close()
    wb.close()
    logging.info("energy change calc done")
    logging.info("submit_job")
    # 5.提交作业进行计算
    create_model(thisTempCae_a, lastTempOdb_b)
    jobName = "Job-Temp-b-{}".format(step + 1)
    myJob = mdb.Job(name=jobName, model='Model-2', numCpus=8, numDomains=8, numGPUs=0)
    myJob.submit()
    myJob.waitForCompletion()
    mdb.saveAs(pathName=TempCae)


# ====================================正式计算============================================
LOG_FORMAT = "%(asctime)s - %(levelname)s - %(message)s"
DATE_FORMAT = "%m/%d/%Y %H:%M:%S %p"
logging.basicConfig(level=logging.INFO, filename='./log.txt', format=LOG_FORMAT, datefmt=DATE_FORMAT)
U = 0.2  # 泊松比
ERRO = 1e-5
TG = 600.0  # 时间间隔单位s
TS = 32
MT = 0  # 全局变量记录单元最大温度
mdb = openMdb(pathName='Model-Temp01.cae')
a1 = mdb.models['Model-1'].rootAssembly.instances['Part-1-1']
NN = len(a1.nodes.getByBoundingBox(-1, -1, 0, 1, 1, 1))  # 全部节点个数
NE = len(a1.elements.getByBoundingBox(-1, -1, 0, 1, 1, 1))  # 全部单元个数
NAN = len(a1.sets['Set-Balls'].nodes)  # 骨料节点个数
NMN = len(a1.sets['Set-Matrix'].nodes)  # 基质节点个数
SN = len(mdb.models['Model-1'].rootAssembly.allSurfaces['Surf-ALL'].nodes)  # 表面节点个数
NME = len(a1.sets['Set-Matrix'].elements)  # 基质单元个数
NAE = len(a1.sets['Set-Balls'].elements)  # 骨料单元个数
# 锁定文件位置信息
# 当前工作目录
work_path = os.getcwd()
# 输出表格的存放目录
excel_path = work_path + "/OutputExcel"
# 输入文件目录
input_dir = "E:/Abaqus/Code/MyCoding/CycleInputFile"
# 各文件位置信息
vapor_table = input_dir + "/VaporTable.xlsx"        # 蒸汽压表格
excel_base = excel_path+"/NNE.xlsx"                 # 初始单元节点信息表
condition_mass = input_dir + "/Condition_mass.txt"  # 水灰比等初始信息
initial_excel = input_dir + "/VertifyData.xlsx"     # 温度场验证文件
CSH, pCSH, CH, AFt, CE, SF, CP, CPF, CP2, CP2D, CPAFt, CPCH, sf, hydc, hydsf, wc, dsf, Em, Eag, \
Sd, Dc0, Dc1, fvAFt, fvCH, fvCSH, inte, ftm, fta = cal_initial(excel_base)
# 开始循环
for step in range(TS):
    elemdh = [[0.0 for i in range(25)] for j in range(NME)]
    elemp = [[0.0 for i in range(21)] for j in range(NME)]
    elempe = [[0.0 for i in range(9)] for j in range(NME)]
    elemvp = [[0.0 for i in range(20)] for j in range(NME)]
    lastTempOdb_a = 'Job-Temp-a-{}.odb'.format(step)
    lastTempOdb_b = 'Job-Temp-b-{}.odb'.format(step)
    lastTempCae_a = 'Model-Temp-a-{}.cae'.format(step)
    lastTempCae_b = 'Model-Temp-b-{}.cae'.format(step)
    thisTempOdb_a = "Job-Temp-a-{}.odb".format(step + 1)
    thisTempOdb_b = "Job-Temp-b-{}.odb".format(step + 1)
    thisTempCae_a = 'Model-Temp-a-{}.cae'.format(step + 1)
    thisTempCae_b = 'Model-Temp-b-{}.cae'.format(step + 1)
    thisMassCae_a = 'Model-Mass-a-{}.cae'.format(step + 1)
    thisMassCae_b = 'Model-Mass-b-{}.cae'.format(step + 1)
    thisMassOdb_a = "Job-Mass-a-{}.odb".format(step + 1)
    thisMassOdb_b = "Job-Mass-b-{}.odb".format(step + 1)
    last_excel_a = excel_path + "/1Step{}.xlsx".format(step)
    this_excel_a = excel_path + "/1Step{}.xlsx".format(step + 1)
    last_excel_b = excel_path + "/2Step{}.xlsx".format(step)
    this_excel_b = excel_path + "/2Step{}.xlsx".format(step + 1)
    if step == 0:
        last_excel_a = this_excel_a
    upgrade_info(False)
    # a.提交作业计算温度
    submit_job_a(lastTempCae_a, lastTempOdb_a, thisTempCae_a)
    # b.读取单步末基质节点温度
    output_temp(this_excel_a, thisTempOdb_a)
    # c.热分解计算
    calc_decomposition(last_excel_a, this_excel_a)
    # d.蒸汽压计算
    calc_vapor(this_excel_a, False)
    # e.质量传输确定
    calc_mass(last_excel_a, this_excel_a, thisMassCae_a, thisMassOdb_a, False)
    # f.再次计算蒸汽压
    calc_vapor(this_excel_a, True)
    # g.输出结点相关数据
    node_info(this_excel_a)
    # ====================考虑吸热后的计算=========================
    upgrade_info(True)
    if step == 0:
        last_excel_b = this_excel_b
    heat_source(this_excel_a, this_excel_b, thisTempCae_b)
    output_temp(this_excel_b, thisTempOdb_b)
    calc_decomposition(last_excel_b, this_excel_b)
    calc_vapor(this_excel_b, False)
    calc_mass(last_excel_b, this_excel_b, thisMassCae_b, thisMassOdb_b, True)
    calc_vapor(this_excel_b, True)
    node_info(this_excel_b)
    if MT >= 600:
        break
    if step == TS - 1:
        logging.info("ALL  DONE!!!")
# done
