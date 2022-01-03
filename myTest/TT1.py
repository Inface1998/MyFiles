# -*- coding: utf-8 -*-
target_dir = "E:/Abaqus/Code/CycleOutputFile/12-30"
m= target_dir+'/step{}/{}-1mm.png'
print (m)
vapor_table = "E:/Abaqus/Code/CycleInputFile/VaporTable.xlsx"

# 等效热导率
mdb.models['Model-1'].materials['Material-Aggregate'].conductivity.setValues(table=((1e-19,),))
# 等效密度
mdb.models['Model-1'].materials['Material-Aggregate'].density.setValues(table=((0.4614,),))
# 等效比容
mdb.models['Model-1'].materials['Material-Aggregate'].specificHeat.setValues(table=((1e2 / (500 + 273.15),),))
# 等效热导率
mdb.models['Model-1'].materials['Material-Matrix'].conductivity.setValues(table=((2.96e-11,),))
# 等效密度
mdb.models['Model-1'].materials['Material-Matrix'].density.setValues(table=((0.4614,),))
# 等效比容
mdb.models['Model-1'].materials['Material-Matrix'].specificHeat.setValues(table=((0.0002253,),))



import openpyxl
a1 = mdb.models['Model-2'].rootAssembly.instances['Part-1-1']
NN = len(a1.nodes.getByBoundingBox(-1, -1, 0, 1, 1, 1))  # 全部节点个数
NE = len(a1.elements.getByBoundingBox(-1, -1, 0, 1, 1, 1))  # 全部单元个数
NAN = len(a1.sets['Set-Balls'].nodes)  # 骨料节点个数
NMN = len(a1.sets['Set-Matrix'].nodes)  # 基质节点个数
SN = len(mdb.models['Model-2'].rootAssembly.allSurfaces['Surf-ALL'].nodes)  # 表面节点个数
NME = len(a1.sets['Set-Matrix'].elements)  # 基质单元个数
NAE = len(a1.sets['Set-Balls'].elements)  # 骨料单元个数
this_excel = "E:/Abaqus/Code/CycleOutputFile/1-3-16/Step25.xlsx"
wb = openpyxl.load_workbook(this_excel)
sheet = wb["MElement"]
sheet2 = wb["AElement"]
a = mdb.models['Model-2'].rootAssembly
c1 = a.instances['Part-1-1'].cells
np = len(c1)
# 为骨料赋予初始蒸气压
vapor_point2 = ()
for n in range(NAE):
    vapor_point2 += ((sheet2.cell(n + 2, 6).value, sheet2.cell(n + 2, 7).value, sheet2.cell(n + 2, 8).value,
                      0.1),)


mdb.models['Model-2'].MappedField(name='AnalyticalField-2',
                                       description='', regionType=POINT, partLevelData=False, localCsys=None,
                                       pointDataFormat=XYZ, fieldDataType=SCALAR, xyzPointData=vapor_point2)
cells2 = c1[0:np - 1]
region2 = regionToolset.Region(cells=cells2)
mdb.models['Model-2'].Temperature(name='Initial-vapor-balls',
                                       createStepName='Initial', region=region2, distributionType=FIELD,
                                       crossSectionDistribution=CONSTANT_THROUGH_THICKNESS,
                                       field='AnalyticalField-2', magnitudes=(1.0,))
vapor_point1 = ()
for i in range(NME):
    vapor_point1 += ((sheet.cell(i + 2, 6).value, sheet.cell(i + 2, 7).value, sheet.cell(i + 2, 8).value,
                      sheet.cell(i + 2, 13).value),)


mdb.models['Model-2'].MappedField(name='AnalyticalField-1',
                                       description='', regionType=POINT, partLevelData=False, localCsys=None,
                                       pointDataFormat=XYZ, fieldDataType=SCALAR, xyzPointData=vapor_point1)
cells1 = c1[np - 1:np]
# 为基质赋予初始蒸气压
region1 = regionToolset.Region(cells=cells1)
mdb.models['Model-2'].Temperature(name='Initial-vapor-matrix',
                                       createStepName='Initial', region=region1, distributionType=FIELD,
                                       crossSectionDistribution=CONSTANT_THROUGH_THICKNESS,
                                       field='AnalyticalField-1', magnitudes=(1.0,))

wb.close()