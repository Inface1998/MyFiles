# -*- coding: UTF-8 -*-    
# 作者：zhangHJ
# 日期：2022/1/4  10:07
import openpyxl


def consider_absorption(excel, n1, n2, label):
    """
    This is a groups style docs.

    Parameters:
      param1 - target excel
      param2 - input column
      param3 - output column
      param4 - output target name
    """

    wb = openpyxl.load_workbook(excel)
    sheet1 = wb["MElement"]
    sheet2 = wb["AllNode"]
    data = [[] for n in range(NN)]
    count = 2
    sheet2.cell(count - 1, n2, label)
    for i in range(2, 101):
        value = sheet1.cell(i, n1).value
        for j in range(2, 6):
            nodeLabel = sheet1.cell(i, j).value
            data[nodeLabel].append(value)
    for item in data:
        if len(item) == 0:
            result = 0.1
        else:
            result = sum(item) / len(item)
        sheet2.cell(count, n2, result)
        count = count + 1
    wb.save(excel)
    wb.close()


NN = 1620
this_excel = "E:/Abaqus/Code/CycleOutputFile/1-5-19/Step{}.xlsx".format(29)
consider_absorption(this_excel, 13, 8, "vapor1")
