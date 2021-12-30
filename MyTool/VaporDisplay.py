# -*- coding: UTF-8 -*-    
# 作者：zhangHJ
# 日期：2021/12/24  09:48
from abaqus import *
from abaqusConstants import *
from caeModules import *
from odbAccess import *
from visualization import *
import os
import openpyxl
import math
import time
import sys
import logging
import numpy as np


def vapor1_display():
    wb = openpyxl.load_workbook(this_excel)
    sheet = wb["MElement"]
    vapor_point = ()
    for i in range(NME):
        vapor_point += ((sheet.cell(i + 2, 6).value, sheet.cell(i + 2, 7).value, sheet.cell(i + 2, 8).value,
                         sheet.cell(i + 2, 13).value),)
    mdb.models['Model-1'].MappedField(name='AnalyticalField-1',
                                      description='', regionType=POINT, partLevelData=False, localCsys=None,
                                      pointDataFormat=XYZ, fieldDataType=SCALAR, xyzPointData=vapor_point)
    a = mdb.models['Model-1'].rootAssembly
    c1 = a.instances['Part-1-1'].cells
    np = len(c1)
    cells1 = c1[np - 1:np]
    # 为基质赋予初始蒸气压
    region = regionToolset.Region(cells=cells1)
    mdb.models['Model-1'].Temperature(name='Vapor-begin',
                                      createStepName='Initial', region=region, distributionType=FIELD,
                                      crossSectionDistribution=CONSTANT_THROUGH_THICKNESS,
                                      field='AnalyticalField-1', magnitudes=(1.0,))
    session.viewports['Viewport: 1'].setValues(displayedObject=session.mdbData['Model-1'])
    session.viewports['Viewport: 1'].odbDisplay.setFrame(step='Initial')
    session.viewports['Viewport: 1'].odbDisplay.setPrimaryVariable(variableLabel='(P) Vapor-begin',
                                                                   outputPosition=NODAL)
    session.viewports['Viewport: 1'].odbDisplay.display.setValues(plotState=(CONTOURS_ON_DEF,))
    session.viewports['Viewport: 1'].viewportAnnotationOptions.setValues(title=OFF, state=OFF)
    session.viewports['Viewport: 1'].view.setValues(session.views['Iso'])
    session.viewports['Viewport: 1'].odbDisplay.viewCuts['X-Plane'].setValues(position=0.01)
    session.pngOptions.setValues(imageSize=(4096, 2057))
    session.viewports['Viewport: 1'].odbDisplay.contourOptions.setValues(maxValue=35, minAutoCompute=OFF, minValue=0.1)
    output_png("begin")


def vapor2_display():
    wb = openpyxl.load_workbook(this_excel)
    sheet = wb["MElement"]
    vapor_point = ()
    for i in range(NME):
        vapor_point += ((sheet.cell(i + 2, 6).value, sheet.cell(i + 2, 7).value, sheet.cell(i + 2, 8).value,
                         sheet.cell(i + 2, 19).value),)
    mdb.models['Model-1'].MappedField(name='AnalyticalField-2',
                                      description='', regionType=POINT, partLevelData=False, localCsys=None,
                                      pointDataFormat=XYZ, fieldDataType=SCALAR, xyzPointData=vapor_point)
    a = mdb.models['Model-1'].rootAssembly
    c1 = a.instances['Part-1-1'].cells
    np = len(c1)
    cells1 = c1[np - 1:np]
    # 为基质赋予初始蒸气压
    region = regionToolset.Region(cells=cells1)
    mdb.models['Model-1'].Temperature(name='Vapor-final',
                                      createStepName='Initial', region=region, distributionType=FIELD,
                                      crossSectionDistribution=CONSTANT_THROUGH_THICKNESS,
                                      field='AnalyticalField-2', magnitudes=(1.0,))
    session.viewports['Viewport: 1'].setValues(displayedObject=session.mdbData['Model-1'])
    session.viewports['Viewport: 1'].odbDisplay.setFrame(step='Initial')
    session.viewports['Viewport: 1'].odbDisplay.setPrimaryVariable(variableLabel='(P) Vapor-final',
                                                                   outputPosition=NODAL)
    session.viewports['Viewport: 1'].odbDisplay.display.setValues(plotState=(CONTOURS_ON_DEF,))
    session.viewports['Viewport: 1'].viewportAnnotationOptions.setValues(title=OFF, state=OFF)
    session.viewports['Viewport: 1'].view.setValues(session.views['Iso'])
    session.viewports['Viewport: 1'].odbDisplay.viewCuts['X-Plane'].setValues(position=0.01)
    session.pngOptions.setValues(imageSize=(4096, 2057))
    session.viewports['Viewport: 1'].odbDisplay.contourOptions.setValues(maxValue=35, minAutoCompute=OFF, minValue=0.1)
    output_png("final")


def output_png(state):
    # 输出文件
    os.mkdir('E:/Abaqus/Code/CycleOutputFile/Display/step{}'.format(i + 1, state))
    session.printToFile(fileName='E:/Abaqus/Code/CycleOutputFile/Display/{}-1mm.png'.format(state), format=PNG,
                        canvasObjects=(session.viewports['Viewport: 1'],))
    session.viewports['Viewport: 1'].odbDisplay.viewCuts['X-Plane'].setValues(position=0.03)
    session.printToFile(fileName='E:/Abaqus/Code/CycleOutputFile/Display/{}-3mm.png'.format(state), format=PNG,
                        canvasObjects=(session.viewports['Viewport: 1'],))
    session.viewports['Viewport: 1'].odbDisplay.viewCuts['X-Plane'].setValues(position=0.05)
    session.printToFile(fileName='E:/Abaqus/Code/CycleOutputFile/Display/{}-5mm.png'.format(state), format=PNG,
                        canvasObjects=(session.viewports['Viewport: 1'],))
    session.viewports['Viewport: 1'].odbDisplay.viewCuts['X-Plane'].setValues(position=0.1)
    session.printToFile(fileName='E:/Abaqus/Code/CycleOutputFile/Display/{}-10mm.png'.format(state), format=PNG,
                        canvasObjects=(session.viewports['Viewport: 1'],))
    leaf = dgo.LeafFromModelElemLabels(elementLabels=get_element())
    session.viewports['Viewport: 1'].odbDisplay.displayGroup.remove(leaf=leaf)


def get_element():
    elements = ()
    for i in range(NAE):
        elements += (('Part-1-1', ('{}'.format(i + 1),)),)
    return elements


files = os.listdir("E:/Abaqus/Code/CycleOutputFile/12-23")
# for i in range(len(files)):
for i in range(1):
    Mdb()
    session.journalOptions.setValues(replayGeometry=INDEX, recoverGeometry=INDEX)
    path_name = "E:/Abaqus/Workpace/Myfile12-30-10/Model-Temp01.cae"
    mdb = openMdb(pathName=path_name)
    a1 = mdb.models['Model-1'].rootAssembly.instances['Part-1-1']
    NN = len(a1.nodes.getByBoundingBox(-1, -1, 0, 1, 1, 1))  # 全部节点个数
    NE = len(a1.elements.getByBoundingBox(-1, -1, 0, 1, 1, 1))  # 全部单元个数
    NAN = len(a1.sets['Set-Balls'].nodes)  # 骨料节点个数
    NMN = len(a1.sets['Set-Matrix'].nodes)  # 基质节点个数
    SN = len(mdb.models['Model-1'].rootAssembly.allSurfaces['Surf-ALL'].nodes)  # 表面节点个数
    NME = len(a1.sets['Set-Matrix'].elements)  # 基质单元个数
    NAE = len(a1.sets['Set-Balls'].elements)  # 骨料单元个数
    this_excel = "E:/Abaqus/Code/CycleOutputFile/12-30/Step{}.xlsx".format(12)
    vapor1_display()
    vapor2_display()
