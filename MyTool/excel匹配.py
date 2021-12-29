# -*- coding: UTF-8 -*-    
# 作者：zhangHJ
# 日期：2021/10/8  16:03
import openpyxl
wb = openpyxl.load_workbook("C:/Users/20454/Desktop/bonus.xlsx")
sheet = wb["Sheet1"]
for i in range(2,48):
     a = sheet.cell(i,7).value
     for j in range(2,48):
          if a == sheet.cell(j,2).value:
               sheet.cell(i,8,sheet.cell(j,3).value)
wb.save("C:/Users/20454/Desktop/bonus.xlsx")
