# -*- coding: UTF-8 -*-    
# 作者：zhangHJ
# 日期：2021/11/9  9:20
import numpy as np
from pylab import *
import openpyxl

# 提取数据
# 创建空数组提取指定时间的温度值
# 确定需要提取的点
data_temp = []
temp_actual = []
data_vapor = []
data_time = []
# 节点编号
nodeNumber = 707
# 总文件数
fn = 31
wb2 = openpyxl.load_workbook('E:/Abaqus/Code/CycleInputFile/VertifyData.xlsx')
st1 = wb2["Sheet1"]
for i in range(1, fn + 1):
    wb = openpyxl.load_workbook('E:/Abaqus/Code/CycleOutputFile/1-14-10/1Step{}.xlsx'.format(i))
    sheet1 = wb["AllNode"]
    data_t = sheet1.cell(nodeNumber + 1, 5).value
    data_temp.append(data_t)
    data_v = sheet1.cell(nodeNumber + 1, 9).value
    if data_v == 0:
        data_v = 0.1
    data_vapor.append(data_v)
    timy = 10 * i
    data_time.append(timy)
    temp_actual.append(st1.cell(4 + 120 * i, 3).value)
    wb.close()
wb2.close()
X = data_time
C = data_temp
S = temp_actual
# 绘制余弦曲线，使用蓝色的、连续的、宽度为 1 （像素）的线条
plot(X, C, color="blue", linewidth=1.0, linestyle="-")
# 绘制正弦曲线，使用绿色的、连续的、宽度为 1 （像素）的线条
plot(X, S, color="green", linewidth=1.0, linestyle="-")

# 设置横轴的上下限,设置横轴记号
xlim(0, 10 * fn)
xticks(np.linspace(0, 10 * fn, fn + 1, endpoint=True))
# 设置纵轴的上下限,设置纵轴记号
ylim(0, 800)
yticks(np.linspace(0, 800, 9, endpoint=True))

# 设置图片边界
xmin, xmax = X[0], X[fn - 1]
ymin, ymax = C[0], C[fn - 1]

dx = (xmax - xmin) * 0.2
dy = (ymax - ymin) * 0.2

xlim(xmin - dx, xmax + dx)
ylim(ymin - dy, ymax + dy)
plt.show()
