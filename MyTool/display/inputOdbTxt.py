# -*- coding: UTF-8 -*-    
# 作者：zhangHJ
# 日期：2022/1/5  12:23
import openpyxl

wb = openpyxl.load_workbook("E:/Abaqus/Code/CycleOutputFile/NNE.xlsx")
with open("E:/Abaqus/Code/CycleInputFile/OdbText.txt", "w+") as f:
    f.write("** odbText.txt\n")
    # 写入Node信息
    f.write("*Node\n")
    sheet1 = wb["AllNode"]
    count = 2
    label = str(sheet1.cell(count, 1).value)
    while True:
        label = str(sheet1.cell(count, 1).value)
        nodeX = str(sheet1.cell(count, 2).value)
        nodeY = str(sheet1.cell(count, 3).value)
        nodeZ = str(sheet1.cell(count, 4).value)
        if label == "None":
            break
        else:
            f.write('%-8s%-18s%-18s%-18s' % (label + ",", nodeX + ",", nodeY + ",", nodeZ) + '\n')
            count = count + 1
    f.write("*MatrixElement\n")
    sheet2 = wb["MElement"]
    count = 2
    label = str(sheet1.cell(count, 1).value)
    while True:
        label = str(sheet2.cell(count, 1).value)
        node1 = str(sheet2.cell(count, 2).value)
        node2 = str(sheet2.cell(count, 3).value)
        node3 = str(sheet2.cell(count, 4).value)
        node4 = str(sheet2.cell(count, 5).value)
        if label == "None":
            break
        else:
            f.write('%-8s%-8s%-8s%-8s%-8s' % (label + ",", node1 + ",", node2 + ",", node3 + ",", node4) + '\n')
            count = count + 1
    f.write("\n")
    f.write("*AggregateElement\n")
    sheet3 = wb["AElement"]
    count = 2
    label = str(sheet1.cell(count, 1).value)
    while True:
        label = str(sheet3.cell(count, 1).value)
        node1 = str(sheet3.cell(count, 2).value)
        node2 = str(sheet3.cell(count, 3).value)
        node3 = str(sheet3.cell(count, 4).value)
        node4 = str(sheet3.cell(count, 5).value)
        if label == "None":
            break
        else:
            f.write('%-8s%-8s%-8s%-8s%-8s' % (label + ",", node1 + ",", node2 + ",", node3 + ",", node4) + '\n')
            count = count + 1
    f.write("\n")
    f.write("*Node matrixVapor\n")
    for i in range(10):
        wb2 = openpyxl.load_workbook("E:/Abaqus/Code/CycleOutputFile/1-13-10/2Step{}.xlsx".format(i + 1))
        f.write("Frame: Increment      0: Step Time = {} \n".format(i + 1))
        s1 = wb2["AllNode"]
        count = 2
        while True:
            label = str(s1.cell(count, 1).value)
            data = str(s1.cell(count, 5).value)
            if label == "None":
                f.write("\n")
                break
            else:
                f.write('%-8s%-18s' % (label + ",", data) + '\n')
                count = count + 1
    wb2.close()
    f.write("\n")
    f.write("*Node aggregateVapor\n")
    for i in range(10):
        wb2 = openpyxl.load_workbook("E:/Abaqus/Code/CycleOutputFile/1-5-19/Step{}.xlsx".format(i + 1))
        f.write("Frame: Increment      0: Step Time = {} \n".format(i + 1))
        s1 = wb2["AllNode"]
        count = 2
        while True:
            label = str(s1.cell(count, 1).value)
            if label == "None":
                f.write("\n")
                break
            else:
                f.write('%-8s%-18s' % (label + ",", 0.1) + '\n')
                count = count + 1
    wb2.close()
wb.close()
